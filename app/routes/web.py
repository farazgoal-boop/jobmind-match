from datetime import datetime, timedelta, timezone
from email.utils import parsedate_to_datetime
import asyncio
import ctypes
import csv
import io
import json
import logging
import os
import platform
import re
import random
import subprocess
import sys
import tempfile
import threading
import time
from collections import deque
from pathlib import Path
from typing import Annotated, Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, PlainTextResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlmodel import Session, select

from app.config import settings
from app.db import get_session
from app.models import AppSetting, ApplicationRecord, CandidateProfile, ClientLead, FilterPreset
from app.services.assisted_apply import (
    build_assisted_links,
    build_client_access_links,
    build_live_search_links,
    build_sales_outreach_links,
)
from app.services.cv_parser import extract_text_from_docx, extract_text_from_pdf
from app.services.matcher import rank_jobs
from app.services.quota import consume_matches, ensure_can_consume_matches
from app.services.source_registry import fetch_jobs_from_sources, free_sources_list

logger = logging.getLogger("jobmind.leadhunter")

router = APIRouter(tags=["web"])
from app.paths import app_root, templates_dir

templates = Jinja2Templates(directory=str(templates_dir()))
templates.env.globals["static_version"] = settings.asset_version
templates.env.globals["app_version"] = settings.app_version
templates.env.globals["app_env"] = settings.app_env

COUNTRY_LIST = [
    "Afghanistan", "Albania", "Algeria", "Argentina", "Armenia", "Australia", "Austria",
    "Azerbaijan", "Bahrain", "Bangladesh", "Belarus", "Belgium", "Bolivia",
    "Bosnia and Herzegovina", "Brazil", "Bulgaria", "Cambodia", "Cameroon", "Canada", "Chile",
    "China", "Colombia", "Costa Rica", "Croatia", "Cuba", "Cyprus", "Czechia", "Denmark",
    "Dominican Republic", "Ecuador", "Egypt", "El Salvador", "Estonia", "Ethiopia", "Finland",
    "France", "Georgia", "Germany", "Ghana", "Greece", "Guatemala", "Honduras", "Hong Kong",
    "Hungary", "Iceland", "India", "Indonesia", "Iran", "Iraq", "Ireland", "Israel", "Italy",
    "Ivory Coast", "Jamaica", "Japan", "Jordan", "Kazakhstan", "Kenya", "Kuwait", "Kyrgyzstan",
    "Laos", "Latvia", "Lebanon", "Libya", "Lithuania", "Luxembourg", "Malaysia", "Maldives",
    "Malta", "Mexico", "Moldova", "Mongolia", "Montenegro", "Morocco", "Myanmar", "Nepal",
    "Netherlands", "New Zealand", "Nicaragua", "Nigeria", "North Macedonia", "Norway", "Oman",
    "Pakistan", "Panama", "Paraguay", "Peru", "Philippines", "Poland", "Portugal", "Qatar",
    "Romania", "Russia", "Rwanda", "Saudi Arabia", "Senegal", "Serbia", "Singapore",
    "Slovakia", "Slovenia", "South Africa", "South Korea", "Spain", "Sri Lanka", "Sudan",
    "Sweden", "Switzerland", "Syria", "Taiwan", "Tajikistan", "Tanzania", "Thailand",
    "Tunisia", "Turkey", "Turkmenistan", "Uganda", "Ukraine", "United Arab Emirates",
    "United Kingdom", "United States", "Uruguay", "Uzbekistan", "Venezuela", "Vietnam",
    "Yemen", "Zambia", "Zimbabwe",
]
templates.env.globals["country_list"] = COUNTRY_LIST

FOLLOW_UP_PATTERN = re.compile(r"\[follow-up:(\d{4}-\d{2}-\d{2})\]")
BUILTIN_RESUME_SOURCES = {
    "final-cv": {
        "label": "Muhammad Faraz Final CV",
        "path": Path(__file__).resolve().parents[2] / "docs" / "final-cv.md",
    }
}

DEFAULT_PROFILE_SEED = {
    "full_name": "Muhammad Faraz",
    "email": "farazgoal@gmail.com",
    "skills_csv": "Python,Flask,FastAPI,JavaScript,SQL,SQLite,HTML,CSS,TailwindCSS",
    "github_username": "farazgoal-boop",
    "resume_url": "",
    "portfolio_url": "https://muhammad-faraz-dev.netlify.app/",
    "linkedin_url": "https://www.linkedin.com/in/m-faraz-85b175179",
    "upwork_url": "https://www.upwork.com/freelancers/~018c67c9c97b482a3a?mp_source=share",
    "fiverr_url": "https://fiverr.com/s/qDKLLXX",
    "product_name": "JobMind Match",
    "product_url": "",
    "sales_pitch": "I build practical Python, FastAPI, Flask, and automation systems for business workflows and software teams.",
}

# Set only in the developer's own local .env, never in .env.example and never
# in a shipped/packaged build's .env -- lets the developer's own dashboard
# start pre-filled with real testing data (name/email/CV text/social links)
# instead of blank on a fresh DB. Mirrors license_service.py's
# JOBMIND_DEV_SKIP_LICENSE. Without this opt-in (the default -- true for
# every customer install, and for the developer's own machine unless
# explicitly set), a fresh profile starts genuinely blank: shipping the
# developer's real identity into a stranger's install is not acceptable for
# a paid product. Also drops the old per-request backfill that silently
# overwrote any blank portfolio_url/linkedin_url/upwork_url/fiverr_url with
# the developer's own links on every dashboard load, even for an existing
# customer profile.
DEV_SEED_PROFILE_ENV_VAR = "JOBMIND_DEV_SEED_PROFILE"

CLIENT_SEARCH_MODES = {"sell_services", "sell_products", "direct_clients"}


def resolve_ui_mode(search_mode: str) -> str:
    return "sell" if search_mode in CLIENT_SEARCH_MODES else "job"


def resolve_dashboard_ui_mode(request: Request, search_mode: str) -> str:
    """Honor ?active_mode=sell|job from mobile APK deep links."""
    query_mode = (request.query_params.get("active_mode") or "").strip().lower()
    if query_mode in {"sell", "job"}:
        return query_mode
    return resolve_ui_mode(search_mode)


def normalize_application_status(status: str) -> str:
    key = (status or "").strip().lower()
    if key == "replied":
        return "offer"
    return key or "saved"


def normalize_client_status(status: str) -> str:
    key = (status or "").strip().lower()
    if key == "new":
        return "discover"
    return key or "discover"


def build_resume_library() -> list[dict[str, str]]:
    library = []
    for key, item in BUILTIN_RESUME_SOURCES.items():
        path = item["path"]
        if path.exists():
            library.append({"key": key, "label": str(item["label"])})
    return library


def read_builtin_resume_text(source_key: str) -> str:
    source = BUILTIN_RESUME_SOURCES.get(source_key)
    if not source:
        raise HTTPException(status_code=404, detail="Resume source not found")

    path = source["path"]
    if not path.exists():
        raise HTTPException(status_code=404, detail="Resume file is missing")

    return path.read_text(encoding="utf-8").strip()


def ensure_seed_profile(session: Session) -> CandidateProfile:
    existing = session.exec(select(CandidateProfile).order_by(CandidateProfile.id.desc())).first()
    if existing:
        return existing

    if os.environ.get(DEV_SEED_PROFILE_ENV_VAR, "").strip() == "1":
        profile = CandidateProfile(
            **DEFAULT_PROFILE_SEED,
            cv_text=read_builtin_resume_text("final-cv") if build_resume_library() else "",
        )
    else:
        profile = CandidateProfile(full_name="", email="")
    session.add(profile)
    session.commit()
    session.refresh(profile)
    return profile


def build_status_summary(applications: list[ApplicationRecord]) -> dict[str, int]:
    summary = {
        "total": len(applications),
        "saved": 0,
        "applied": 0,
        "interview": 0,
        "offer": 0,
        "rejected": 0,
    }
    for row in applications:
        key = normalize_application_status(row.status)
        if key in summary:
            summary[key] += 1
    return summary


def build_company_summary(applications: list[ApplicationRecord]) -> list[dict[str, int | str]]:
    counts: dict[str, dict[str, int | str]] = {}
    for row in applications:
        bucket = counts.setdefault(
            row.company,
            {"company": row.company, "total": 0, "applied": 0, "interview": 0, "offer": 0},
        )
        bucket["total"] += 1
        status = normalize_application_status(row.status)
        if status in {"applied", "interview", "offer"}:
            bucket[status] += 1
    return sorted(counts.values(), key=lambda item: (-int(item["total"]), str(item["company"])))[:6]


def extract_follow_up_date(notes: str) -> str:
    match = FOLLOW_UP_PATTERN.search(notes or "")
    return match.group(1) if match else ""


def strip_follow_up_marker(notes: str) -> str:
    cleaned = FOLLOW_UP_PATTERN.sub("", notes or "")
    return cleaned.strip(" |")


def compose_notes(notes: str, follow_up_date: str) -> str:
    cleaned = strip_follow_up_marker(notes)
    marker = f"[follow-up:{follow_up_date}]" if follow_up_date else ""
    return " | ".join(part for part in [marker, cleaned] if part)


def build_tracker_entries(applications: list[ApplicationRecord]) -> list[dict]:
    entries = []
    for row in applications:
        entries.append(
            {
                "id": row.id,
                "candidate_id": row.candidate_id,
                "job_title": row.job_title,
                "company": row.company,
                "source": row.source,
                "job_url": row.job_url,
                "status": normalize_application_status(row.status),
                "notes": strip_follow_up_marker(row.notes),
                "follow_up_date": extract_follow_up_date(row.notes),
                "updated_at": row.updated_at.isoformat(),
            }
        )
    return entries


def build_follow_up_summary(applications: list[ApplicationRecord]) -> list[dict[str, str]]:
    items = []
    for row in applications:
        follow_up_date = extract_follow_up_date(row.notes)
        if not follow_up_date:
            continue
        items.append(
            {
                "company": row.company,
                "job_title": row.job_title,
                "follow_up_date": follow_up_date,
                "status": row.status,
            }
        )
    return sorted(items, key=lambda item: item["follow_up_date"])[:6]


def build_client_status_summary(leads: list[ClientLead]) -> dict[str, int]:
    summary = {
        "total": len(leads),
        "discover": 0,
        "contacted": 0,
        "replied": 0,
        "negotiation": 0,
        "won": 0,
        "lost": 0,
    }
    for row in leads:
        key = normalize_client_status(row.status)
        if key in summary:
            summary[key] += 1
    return summary


def build_client_lead_entries(leads: list[ClientLead]) -> list[dict]:
    return [
        {
            "id": row.id,
            "candidate_id": row.candidate_id,
            "lead_name": row.lead_name,
            "company": row.company,
            "source": row.source,
            "lead_url": row.lead_url,
            "contact_email": row.contact_email,
            "contact_phone": row.contact_phone,
            "contact_channel": row.contact_channel,
            "offer_type": row.offer_type,
            "status": normalize_client_status(row.status),
            "notes": row.notes,
            "follow_up_date": row.follow_up_date,
            "location": row.location,
            "updated_at": row.updated_at.isoformat(),
        }
        for row in leads
    ]


def build_filter_preset_entries(presets: list[FilterPreset]) -> list[dict[str, str | int]]:
    return [
        {
            "id": row.id,
            "candidate_id": row.candidate_id,
            "name": row.name,
            "preset_json": row.preset_json,
            "updated_at": row.updated_at.isoformat(),
        }
        for row in presets
    ]


def build_client_follow_up_summary(leads: list[ClientLead]) -> list[dict[str, str]]:
    items = []
    for row in leads:
        if not row.follow_up_date:
            continue
        items.append(
            {
                "company": row.company,
                "lead_name": row.lead_name or "Decision maker",
                "follow_up_date": row.follow_up_date,
                "status": row.status,
            }
        )
    return sorted(items, key=lambda item: item["follow_up_date"])[:6]


def build_outreach_links(profile: CandidateProfile | None, offer_type: str) -> list[dict[str, str]]:
    if not profile:
        return []
    return build_sales_outreach_links(
        full_name=profile.full_name,
        email=profile.email,
        offer_type=offer_type,
        product_name=profile.product_name,
        product_url=profile.product_url,
        resume_url=profile.resume_url,
        portfolio_url=profile.portfolio_url,
        sales_pitch=profile.sales_pitch,
    )


def normalize_platform_targets(platform_targets: str) -> list[str]:
    selected_targets = [target.strip().lower() for target in platform_targets.split(",") if target.strip()]
    return selected_targets or ["indeed", "linkedin", "upwork", "google_clients"]


def resolve_selected_sources(request: Request, sources: str, is_premium: bool) -> list[str]:
    source_tokens = [source.strip().lower() for source in sources.split(",") if source.strip()]
    if "sources" in request.query_params and not source_tokens:
        return []

    selected_sources = source_tokens or free_sources_list()
    if not is_premium:
        selected_sources = selected_sources[:2]
    return selected_sources


def is_external_only_mode(search_mode: str, selected_sources: list[str], selected_platform_targets: list[str]) -> bool:
    external_job_targets = {"upwork", "fiverr", "linkedin"}
    return search_mode == "job_search" and not selected_sources and any(
        target in external_job_targets for target in selected_platform_targets
    )


def is_client_only_mode(search_mode: str, selected_sources: list[str], selected_platform_targets: list[str]) -> bool:
    client_modes = {"sell_services", "sell_products", "direct_clients"}
    return search_mode in client_modes and not selected_sources and bool(selected_platform_targets)


def score_to_level(score: float) -> str:
    if score >= 0.65:
        return "high"
    if score >= 0.45:
        return "medium"
    return "low"


def is_backend_job(job: dict) -> bool:
    text = f"{job.get('title', '')} {job.get('description', '')}".lower()
    backend_terms = ["python", "backend", "api", "fastapi", "django", "flask", "automation", "scrap"]
    frontend_terms = ["react", "frontend", "front end", "vue", "angular", "ui designer", "css"]
    has_backend = any(term in text for term in backend_terms)
    has_frontend = any(term in text for term in frontend_terms)
    return has_backend and not has_frontend


def is_remote_job(job: dict) -> bool:
    text = f"{job.get('title', '')} {job.get('description', '')} {job.get('location', '')}".lower()
    return any(term in text for term in ["remote", "work from home", "distributed", "anywhere"])


def is_junior_friendly_job(job: dict) -> bool:
    text = f"{job.get('title', '')} {job.get('description', '')}".lower()
    positive = ["junior", "entry", "associate", "graduate", "trainee", "intern"]
    negative = ["senior", "staff", "principal", "lead", "manager", "director"]
    return any(term in text for term in positive) or not any(term in text for term in negative)


def is_pakistan_friendly_job(job: dict) -> bool:
    text = f"{job.get('title', '')} {job.get('description', '')} {job.get('location', '')}".lower()
    blocking_terms = ["us only", "usa only", "canada only", "europe only", "eu only", "uk only"]
    asia_terms = ["pakistan", "asia", "worldwide", "global", "anywhere", "remote"]
    return not any(term in text for term in blocking_terms) and any(term in text for term in asia_terms)


def has_salary_signal(job: dict) -> bool:
    salary = str(job.get("salary", "") or "").strip()
    text = f"{salary} {job.get('description', '')}".lower()
    return bool(salary) or any(term in text for term in ["salary", "compensation", "$", "usd", "eur", "per year"])


def has_visa_signal(job: dict) -> bool:
    if job.get("visa_support") is True:
        return True
    text = f"{job.get('description', '')} {job.get('location', '')}".lower()
    return any(term in text for term in ["visa", "sponsorship", "relocation"])


def matches_region(job: dict, region: str) -> bool:
    if region == "any":
        return True
    text = f"{job.get('location', '')} {job.get('description', '')}".lower()
    region_terms = {
        "pakistan": ["pakistan"],
        "asia": ["asia", "singapore", "india", "pakistan", "uae", "japan", "malaysia"],
        "europe": ["europe", "eu", "germany", "france", "spain", "italy", "netherlands", "poland"],
        "americas": ["america", "americas", "usa", "united states", "canada", "latam", "brazil", "mexico"],
        "global": ["worldwide", "global", "anywhere", "remote"],
    }
    return any(term in text for term in region_terms.get(region, []))


def matches_search_focus(job: dict, search_focus: str) -> bool:
    if not search_focus or search_focus == "all":
        return True

    text = f"{job.get('title', '')} {job.get('description', '')}".lower()
    focus_terms = {
        "python": ["python", "django", "flask", "fastapi"],
        "fastapi": ["fastapi"],
        "flask": ["flask"],
        "fullstack": ["full stack", "full-stack", "backend", "frontend"],
    }
    return any(term in text for term in focus_terms.get(search_focus, []))


def parse_job_datetime(raw_value: str) -> datetime | None:
    raw_text = str(raw_value or "").strip()
    if not raw_text:
        return None

    try:
        normalized = raw_text.replace("Z", "+00:00")
        parsed = datetime.fromisoformat(normalized)
    except ValueError:
        try:
            parsed = parsedate_to_datetime(raw_text)
        except (TypeError, ValueError):
            return None

    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def matches_posted_window(job: dict, posted_within: str) -> bool:
    if posted_within in {"", "any"}:
        return True

    published_at = parse_job_datetime(job.get("published_at", ""))
    if published_at is None:
        return False

    day_windows = {
        "1d": 1,
        "3d": 3,
        "7d": 7,
        "14d": 14,
        "30d": 30,
    }
    max_age_days = day_windows.get(posted_within)
    if max_age_days is None:
        return True

    return published_at >= datetime.now(timezone.utc) - timedelta(days=max_age_days)


def build_match_summary(matches: list[dict]) -> dict[str, int]:
    summary = {"total": len(matches), "high": 0, "medium": 0, "low": 0, "remote": 0, "salary": 0, "visa": 0}
    for item in matches:
        level = item.get("match_level", "low")
        if level in summary:
            summary[level] += 1
        if is_remote_job(item.get("job", {})):
            summary["remote"] += 1
        if item.get("has_salary"):
            summary["salary"] += 1
        if item.get("has_visa_support"):
            summary["visa"] += 1
    return summary


def compute_fit_score(score: float) -> int:
    return max(0, min(100, round(score * 100)))


def compute_trust_score(job: dict) -> int:
    score = 40
    source = str(job.get("source", "") or "").lower()
    description = str(job.get("description", "") or "").lower()
    location = str(job.get("location", "") or "").lower()
    url = str(job.get("url", "") or "").lower()

    trusted_sources = {
        "remotive": 18,
        "weworkremotely": 16,
        "arbeitnow": 14,
    }
    score += trusted_sources.get(source, 8)

    if has_salary_signal(job):
        score += 10
    if has_visa_signal(job):
        score += 6
    if is_remote_job(job):
        score += 6
    if any(token in description for token in ["apply", "benefits", "team", "requirements"]):
        score += 8
    if any(token in location for token in ["remote", "worldwide", "global"]):
        score += 4
    if url.startswith("https://"):
        score += 4

    return max(0, min(100, score))


def build_match_reasons(job: dict, matched_skills: list[str], missing_skills: list[str]) -> list[str]:
    reasons: list[str] = []
    if matched_skills:
        reasons.append(f"Stack match: {', '.join(matched_skills[:3])}")
    if is_remote_job(job):
        reasons.append("Remote-ready listing")
    if is_junior_friendly_job(job):
        reasons.append("Junior-friendly wording")
    if has_salary_signal(job):
        reasons.append("Compensation visible")
    if has_visa_signal(job):
        reasons.append("Visa or relocation signal")
    if is_pakistan_friendly_job(job):
        reasons.append("Pakistan-friendly access")
    if missing_skills:
        reasons.append(f"Upskill next: {', '.join(missing_skills[:2])}")
    return reasons[:5]


def build_active_filter_badges(
    search_mode: str,
    offer_type: str,
    client_type: str,
    contact_goal: str,
    posted_within: str,
    counterparty_type: str,
    trust_signal: str,
    company_size: str,
    proposal_pressure: str,
    verified_payment_only: bool,
) -> list[str]:
    badges = [
        f"Mode: {search_mode.replace('_', ' ')}",
        f"Offer: {offer_type.replace('_', ' ')}",
        f"Client: {client_type.replace('_', ' ')}",
        f"Goal: {contact_goal.replace('_', ' ')}",
    ]
    if posted_within != "any":
        badges.append(f"Posted: {posted_within}")
    if counterparty_type != "any":
        badges.append(f"Side: {counterparty_type.replace('_', ' ')}")
    if trust_signal != "any":
        badges.append(f"Trust: {trust_signal.replace('_', ' ')}")
    if company_size != "any":
        badges.append(f"Size: {company_size.replace('_', ' ')}")
    if proposal_pressure != "any":
        badges.append(f"Competition: {proposal_pressure}")
    if verified_payment_only:
        badges.append("Verified payment")
    return badges


def build_dashboard_live_links(
    profile: CandidateProfile | None,
    search_mode: str,
    offer_type: str,
    client_type: str,
    demand_level: str,
    contact_goal: str,
    counterparty_type: str,
    posted_within: str,
    verified_payment_only: bool,
    trust_signal: str,
    company_size: str,
    proposal_pressure: str,
    platform_targets: str,
    search_focus: str,
    region_preference: str,
    remote_only: bool,
    junior_only: bool,
    backend_only: bool,
    pakistan_friendly_only: bool,
    salary_only: bool,
    visa_support_only: bool,
    custom_keywords: str,
) -> list[dict[str, str]]:
    if not profile:
        return []

    return build_live_search_links(
        search_mode=search_mode,
        offer_type=offer_type,
        client_type=client_type,
        demand_level=demand_level,
        contact_goal=contact_goal,
        counterparty_type=counterparty_type,
        posted_within=posted_within,
        verified_payment_only=verified_payment_only,
        trust_signal=trust_signal,
        company_size=company_size,
        proposal_pressure=proposal_pressure,
        platform_targets=[target.strip() for target in platform_targets.split(",") if target.strip()],
        search_focus=search_focus,
        region_preference=region_preference,
        remote_only=remote_only,
        junior_only=junior_only,
        backend_only=backend_only,
        pakistan_friendly_only=pakistan_friendly_only,
        salary_only=salary_only,
        visa_support_only=visa_support_only,
        candidate_skills=profile.skills_csv,
        custom_keywords=custom_keywords,
    )


def build_dashboard_client_links(
    profile: CandidateProfile | None,
    search_mode: str,
    offer_type: str,
    client_type: str,
    counterparty_type: str,
    region_preference: str,
    contact_goal: str,
    posted_within: str,
    verified_payment_only: bool,
    trust_signal: str,
    company_size: str,
    custom_keywords: str,
) -> list[dict[str, str]]:
    if not profile:
        return []

    return build_client_access_links(
        search_mode=search_mode,
        offer_type=offer_type,
        client_type=client_type,
        counterparty_type=counterparty_type,
        region_preference=region_preference,
        contact_goal=contact_goal,
        posted_within=posted_within,
        verified_payment_only=verified_payment_only,
        trust_signal=trust_signal,
        company_size=company_size,
        custom_keywords=custom_keywords,
    )


def build_client_search_results(
    live_search_links: list[dict[str, str]],
    client_access_links: list[dict[str, str]],
    offer_type: str,
    client_type: str,
    contact_goal: str,
    counterparty_type: str,
    custom_keywords: str,
    top_k: int,
) -> list[dict[str, str]]:
    seen_urls: set[str] = set()
    query_focus = custom_keywords.strip() or client_type.replace("_", " ")
    target_label = counterparty_type.replace("_", " ") if counterparty_type != "any" else "decision maker"
    offer_label = offer_type.replace("_", " ")

    results: list[dict[str, str]] = []
    for link in [*live_search_links, *client_access_links]:
        url = str(link.get("url", "") or "").strip()
        platform = str(link.get("platform", "") or "Client Route").strip()
        caption = str(link.get("caption", "") or "Open buyer route").strip()
        if not url or url in seen_urls:
            continue
        seen_urls.add(url)

        source_label = "Search Route" if link in live_search_links else "Client Access"
        company = f"{platform} {client_type.replace('_', ' ')} prospects".title()
        title = f"{platform} {offer_label.title()} buyers"
        notes = (
            f"Target {target_label} for {query_focus}. "
            f"Goal: {contact_goal.replace('_', ' ')}. "
            f"Route: {caption}."
        )
        results.append(
            {
                "title": title,
                "company": company,
                "platform": platform,
                "caption": caption,
                "url": url,
                "source": source_label,
                "query_focus": query_focus,
                "target_label": target_label,
                "notes": notes,
                "contact_channel": "linkedin" if "linkedin" in platform.lower() else "website",
                "lead_name": target_label.title(),
                "offer_type": offer_type,
                "follow_up_hint": "Open route, shortlist a real buyer, then save the lead below.",
            }
        )
        if len(results) >= top_k:
            break
    return results


@router.get("/dashboard", response_class=HTMLResponse)
def dashboard(request: Request, session: Annotated[Session, Depends(get_session)], candidate_id: int | None = None):
    seed_profile = ensure_seed_profile(session)
    profiles = session.exec(select(CandidateProfile).order_by(CandidateProfile.id.desc())).all()
    selected = session.get(CandidateProfile, candidate_id) if candidate_id else seed_profile
    selected_platform_targets = normalize_platform_targets("indeed,linkedin")
    resume_library = build_resume_library()
    applications = []
    client_leads = []
    filter_presets = []
    if selected:
        applications = session.exec(
            select(ApplicationRecord)
            .where(ApplicationRecord.candidate_id == selected.id)
            .order_by(ApplicationRecord.updated_at.desc())
        ).all()
        client_leads = session.exec(
            select(ClientLead).where(ClientLead.candidate_id == selected.id).order_by(ClientLead.updated_at.desc())
        ).all()
        filter_presets = session.exec(
            select(FilterPreset).where(FilterPreset.candidate_id == selected.id).order_by(FilterPreset.name.asc())
        ).all()
    status_summary = build_status_summary(applications)
    client_status_summary = build_client_status_summary(client_leads)
    live_search_links = build_dashboard_live_links(
        profile=selected,
        search_mode="job_search",
        offer_type="software",
        client_type="startup",
        demand_level="latest",
        contact_goal="apply",
        counterparty_type="any",
        posted_within="7d",
        verified_payment_only=False,
        trust_signal="verified_company",
        company_size="any",
        proposal_pressure="any",
        platform_targets="indeed,linkedin",
        search_focus="python",
        region_preference="global",
        remote_only=True,
        junior_only=False,
        backend_only=True,
        pakistan_friendly_only=False,
        salary_only=False,
        visa_support_only=False,
        custom_keywords="",
    )
    client_access_links = build_dashboard_client_links(
        profile=selected,
        search_mode="direct_clients",
        offer_type="software",
        client_type="startup",
        counterparty_type="any",
        region_preference="global",
        contact_goal="pitch",
        posted_within="7d",
        verified_payment_only=False,
        trust_signal="any",
        company_size="any",
        custom_keywords="",
    )
    active_filter_badges = build_active_filter_badges(
        search_mode="job_search",
        offer_type="software",
        client_type="startup",
        contact_goal="apply",
        posted_within="7d",
        counterparty_type="any",
        trust_signal="any",
        company_size="any",
        proposal_pressure="any",
        verified_payment_only=False,
    )
    outreach_links = build_outreach_links(selected, "software")
    active_ui_mode = resolve_dashboard_ui_mode(request, "job_search")
    return templates.TemplateResponse(
        "dashboard.html",
        {
            "request": request,
            "profiles": profiles,
            "selected": selected,
            "matches": [],
            "usage": None,
            "error": "",
            "applications": build_tracker_entries(applications),
            "client_leads": build_client_lead_entries(client_leads),
            "saved_filter_presets": build_filter_preset_entries(filter_presets),
            "status_summary": status_summary,
            "client_status_summary": client_status_summary,
            "company_summary": build_company_summary(applications),
            "follow_up_summary": build_follow_up_summary(applications),
            "client_follow_up_summary": build_client_follow_up_summary(client_leads),
            "live_search_links": live_search_links,
            "client_access_links": client_access_links,
            "outreach_links": outreach_links,
            "resume_library": resume_library,
            "sources": "remotive,weworkremotely,arbeitnow",
            "platform_targets": "indeed,linkedin",
            "selected_platform_targets": selected_platform_targets,
            "external_only_mode": False,
            "client_only_mode": False,
            "search_mode": "job_search",
            "offer_type": "software",
            "client_type": "startup",
            "demand_level": "latest",
            "contact_goal": "apply",
            "counterparty_type": "any",
            "posted_within": "7d",
            "verified_payment_only": False,
            "trust_signal": "verified_company",
            "company_size": "any",
            "proposal_pressure": "any",
            "custom_keywords": "",
            "min_match_level": "all",
            "backend_only": True,
            "remote_only": True,
            "junior_only": False,
            "search_focus": "python",
            "pakistan_friendly_only": False,
            "salary_only": False,
            "visa_support_only": False,
            "region_preference": "global",
            "active_ui_mode": active_ui_mode,
            "match_summary": build_match_summary([]),
            "active_filter_badges": active_filter_badges,
            "client_search_results": [],
        },
    )


@router.post("/dashboard/applications")
def add_application_from_form(
    candidate_id: Annotated[int, Form(...)],
    job_title: Annotated[str, Form(...)],
    company: Annotated[str, Form(...)],
    source: Annotated[str, Form(...)],
    job_url: Annotated[str, Form(...)],
    status: Annotated[str, Form()] = "saved",
    notes: Annotated[str, Form()] = "",
    follow_up_date: Annotated[str, Form()] = "",
    session: Annotated[Session, Depends(get_session)] = None,
):
    row = ApplicationRecord(
        candidate_id=candidate_id,
        job_title=job_title,
        company=company,
        source=source,
        job_url=job_url,
        status=normalize_application_status(status),
        notes=compose_notes(notes, follow_up_date),
    )
    session.add(row)
    session.commit()
    return RedirectResponse(url=f"/dashboard?candidate_id={candidate_id}", status_code=303)


@router.post("/dashboard/applications/{application_id}")
def update_application_from_form(
    application_id: int,
    candidate_id: Annotated[int, Form(...)],
    status: Annotated[str, Form(...)],
    notes: Annotated[str, Form()] = "",
    follow_up_date: Annotated[str, Form()] = "",
    session: Annotated[Session, Depends(get_session)] = None,
):
    row = session.get(ApplicationRecord, application_id)
    if not row:
        raise HTTPException(status_code=404, detail="Application not found")

    row.status = normalize_application_status(status)
    row.notes = compose_notes(notes, follow_up_date)
    row.updated_at = datetime.utcnow()
    session.add(row)
    session.commit()
    return RedirectResponse(url=f"/dashboard?candidate_id={candidate_id}", status_code=303)


@router.post("/dashboard/profile")
def create_profile_from_form(
    full_name: Annotated[str, Form(...)],
    email: Annotated[str, Form(...)],
    skills_csv: Annotated[str, Form(...)],
    profile_id: Annotated[str | None, Form()] = None,
    github_username: Annotated[str, Form()] = "",
    resume_url: Annotated[str, Form()] = "",
    portfolio_url: Annotated[str, Form()] = "",
    linkedin_url: Annotated[str, Form()] = "",
    upwork_url: Annotated[str, Form()] = "",
    fiverr_url: Annotated[str, Form()] = "",
    product_name: Annotated[str, Form()] = "",
    product_url: Annotated[str, Form()] = "",
    sales_pitch: Annotated[str, Form()] = "",
    session: Annotated[Session, Depends(get_session)] = None,
):
    normalized_profile_id = int(profile_id) if profile_id and profile_id.strip() else None
    profile = session.get(CandidateProfile, normalized_profile_id) if normalized_profile_id else CandidateProfile()
    profile.full_name = full_name
    profile.email = email
    profile.skills_csv = skills_csv
    profile.github_username = github_username
    profile.resume_url = resume_url
    profile.portfolio_url = portfolio_url
    profile.linkedin_url = linkedin_url
    profile.upwork_url = upwork_url
    profile.fiverr_url = fiverr_url
    profile.product_name = product_name
    profile.product_url = product_url
    profile.sales_pitch = sales_pitch
    session.add(profile)
    session.commit()
    session.refresh(profile)
    return RedirectResponse(url=f"/dashboard?candidate_id={profile.id}", status_code=303)


@router.post("/dashboard/client-leads")
def add_client_lead_from_form(
    candidate_id: Annotated[int, Form(...)],
    company: Annotated[str, Form(...)],
    lead_name: Annotated[str, Form()] = "",
    source: Annotated[str, Form()] = "manual",
    lead_url: Annotated[str, Form()] = "",
    contact_email: Annotated[str, Form()] = "",
    contact_phone: Annotated[str, Form()] = "",
    contact_channel: Annotated[str, Form()] = "email",
    offer_type: Annotated[str, Form()] = "services",
    status: Annotated[str, Form()] = "discover",
    notes: Annotated[str, Form()] = "",
    follow_up_date: Annotated[str, Form()] = "",
    session: Annotated[Session, Depends(get_session)] = None,
):
    row = ClientLead(
        candidate_id=candidate_id,
        company=company,
        lead_name=lead_name,
        source=source,
        lead_url=lead_url,
        contact_email=contact_email,
        contact_phone=contact_phone,
        contact_channel=contact_channel,
        offer_type=offer_type,
        status=normalize_client_status(status),
        notes=notes,
        follow_up_date=follow_up_date,
    )
    session.add(row)
    session.commit()
    return RedirectResponse(url=f"/dashboard?candidate_id={candidate_id}", status_code=303)


@router.post("/dashboard/client-leads/{lead_id}")
def update_client_lead_from_form(
    lead_id: int,
    candidate_id: Annotated[int, Form(...)],
    status: Annotated[str, Form(...)],
    notes: Annotated[str, Form()] = "",
    follow_up_date: Annotated[str, Form()] = "",
    contact_email: Annotated[str, Form()] = "",
    contact_phone: Annotated[str, Form()] = "",
    contact_channel: Annotated[str, Form()] = "email",
    session: Annotated[Session, Depends(get_session)] = None,
):
    row = session.get(ClientLead, lead_id)
    if not row:
        raise HTTPException(status_code=404, detail="Client lead not found")

    row.status = normalize_client_status(status)
    row.notes = notes
    row.follow_up_date = follow_up_date
    row.contact_email = contact_email
    row.contact_phone = contact_phone
    row.contact_channel = contact_channel
    row.updated_at = datetime.utcnow()
    session.add(row)
    session.commit()
    return RedirectResponse(url=f"/dashboard?candidate_id={candidate_id}", status_code=303)


@router.post("/dashboard/filter-presets")
def save_filter_preset(
    candidate_id: Annotated[int, Form(...)],
    preset_name: Annotated[str, Form(...)],
    preset_payload: Annotated[str, Form(...)],
    session: Annotated[Session, Depends(get_session)] = None,
):
    candidate = session.get(CandidateProfile, candidate_id)
    if not candidate:
        raise HTTPException(status_code=404, detail="Candidate not found")

    name = preset_name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Preset name required")

    try:
        payload = json.loads(preset_payload)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=400, detail="Invalid preset payload") from exc

    row = session.exec(
        select(FilterPreset).where(FilterPreset.candidate_id == candidate_id, FilterPreset.name == name)
    ).first()
    if row:
        row.preset_json = json.dumps(payload)
        row.updated_at = datetime.utcnow()
    else:
        row = FilterPreset(candidate_id=candidate_id, name=name, preset_json=json.dumps(payload))

    session.add(row)
    session.commit()
    session.refresh(row)
    return JSONResponse({"ok": True, "preset": build_filter_preset_entries([row])[0]})


@router.post("/dashboard/filter-presets/{preset_id}/delete")
def delete_filter_preset(
    preset_id: int,
    candidate_id: Annotated[int, Form(...)],
    session: Annotated[Session, Depends(get_session)] = None,
):
    row = session.get(FilterPreset, preset_id)
    if not row or row.candidate_id != candidate_id:
        raise HTTPException(status_code=404, detail="Preset not found")

    session.delete(row)
    session.commit()
    return JSONResponse({"ok": True, "preset_id": preset_id})


@router.post("/dashboard/upload-cv")
async def upload_cv_from_form(
    candidate_id: Annotated[int, Form(...)],
    file: UploadFile = File(...),
    session: Annotated[Session, Depends(get_session)] = None,
):
    profile = session.get(CandidateProfile, candidate_id)
    if not profile:
        raise HTTPException(status_code=404, detail="Candidate not found")

    content = await file.read()
    filename = (file.filename or "").lower()
    if filename.endswith(".pdf"):
        cv_text = extract_text_from_pdf(content)
    elif filename.endswith(".docx"):
        cv_text = extract_text_from_docx(content)
    else:
        raise HTTPException(status_code=400, detail="Only PDF or DOCX supported")

    profile.cv_text = cv_text
    session.add(profile)
    session.commit()
    return RedirectResponse(url=f"/dashboard?candidate_id={profile.id}", status_code=303)


@router.post("/dashboard/import-cv")
def import_cv_from_library(
    candidate_id: Annotated[int, Form(...)],
    source_key: Annotated[str, Form(...)],
    session: Annotated[Session, Depends(get_session)] = None,
):
    profile = session.get(CandidateProfile, candidate_id)
    if not profile:
        raise HTTPException(status_code=404, detail="Candidate not found")

    source_key = source_key.strip()
    if not source_key:
        raise HTTPException(status_code=400, detail="Choose a saved resume before importing")

    profile.cv_text = read_builtin_resume_text(source_key)
    session.add(profile)
    session.commit()
    return RedirectResponse(url=f"/dashboard?candidate_id={profile.id}", status_code=303)


@router.get("/dashboard/matches", response_class=HTMLResponse)
def dashboard_matches(
    request: Request,
    candidate_id: int,
    top_k: int = 5,
    sources: str = "",
    platform_targets: str = "indeed,linkedin",
    search_mode: str = "job_search",
    offer_type: str = "software",
    client_type: str = "startup",
    demand_level: str = "latest",
    contact_goal: str = "apply",
    counterparty_type: str = "any",
    posted_within: str = "7d",
    verified_payment_only: bool = False,
    trust_signal: str = "verified_company",
    company_size: str = "any",
    proposal_pressure: str = "any",
    custom_keywords: str = "",
    min_match_level: str = "all",
    backend_only: bool = True,
    remote_only: bool = True,
    junior_only: bool = False,
    search_focus: str = "python",
    pakistan_friendly_only: bool = False,
    salary_only: bool = False,
    visa_support_only: bool = False,
    region_preference: str = "global",
    session: Annotated[Session, Depends(get_session)] = None,
):
    profiles = session.exec(select(CandidateProfile).order_by(CandidateProfile.id.desc())).all()
    selected = session.get(CandidateProfile, candidate_id)
    selected_platform_targets = normalize_platform_targets(platform_targets)
    resume_library = build_resume_library()
    if not selected:
        return templates.TemplateResponse(
            "dashboard.html",
            {
                "request": request,
                "profiles": profiles,
                "selected": None,
                "matches": [],
                "usage": None,
                "error": "Candidate not found.",
                "applications": [],
                "client_leads": [],
                "status_summary": build_status_summary([]),
                "client_status_summary": build_client_status_summary([]),
                "company_summary": [],
                "follow_up_summary": [],
                "client_follow_up_summary": [],
                "live_search_links": [],
                "client_access_links": [],
                "outreach_links": [],
                "resume_library": resume_library,
                "sources": ",".join(free_sources_list()),
                "platform_targets": platform_targets,
                "selected_platform_targets": selected_platform_targets,
                "external_only_mode": False,
                "client_only_mode": False,
                "search_mode": search_mode,
                "offer_type": offer_type,
                "client_type": client_type,
                "demand_level": demand_level,
                "contact_goal": contact_goal,
                "counterparty_type": counterparty_type,
                "posted_within": posted_within,
                "verified_payment_only": verified_payment_only,
                "trust_signal": trust_signal,
                "company_size": company_size,
                "proposal_pressure": proposal_pressure,
                "custom_keywords": custom_keywords,
                "min_match_level": min_match_level,
                "backend_only": backend_only,
                "remote_only": remote_only,
                "junior_only": junior_only,
                "search_focus": search_focus,
                "pakistan_friendly_only": pakistan_friendly_only,
                "salary_only": salary_only,
                "visa_support_only": visa_support_only,
                "region_preference": region_preference,
                "match_summary": build_match_summary([]),
                "client_search_results": [],
            },
        )

    applications = session.exec(
        select(ApplicationRecord)
        .where(ApplicationRecord.candidate_id == selected.id)
        .order_by(ApplicationRecord.updated_at.desc())
    ).all()
    client_leads = session.exec(
        select(ClientLead).where(ClientLead.candidate_id == selected.id).order_by(ClientLead.updated_at.desc())
    ).all()
    filter_presets = session.exec(
        select(FilterPreset).where(FilterPreset.candidate_id == selected.id).order_by(FilterPreset.name.asc())
    ).all()
    status_summary = build_status_summary(applications)
    client_status_summary = build_client_status_summary(client_leads)

    selected_sources = resolve_selected_sources(request, sources, selected.is_premium)
    external_only_mode = is_external_only_mode(search_mode, selected_sources, selected_platform_targets)
    client_only_mode = is_client_only_mode(search_mode, selected_sources, selected_platform_targets)
    active_ui_mode = resolve_dashboard_ui_mode(request, search_mode)

    live_search_links = build_dashboard_live_links(
        profile=selected,
        search_mode=search_mode,
        offer_type=offer_type,
        client_type=client_type,
        demand_level=demand_level,
        contact_goal=contact_goal,
        counterparty_type=counterparty_type,
        posted_within=posted_within,
        verified_payment_only=verified_payment_only,
        trust_signal=trust_signal,
        company_size=company_size,
        proposal_pressure=proposal_pressure,
        platform_targets=platform_targets,
        search_focus=search_focus,
        region_preference=region_preference,
        remote_only=remote_only,
        junior_only=junior_only,
        backend_only=backend_only,
        pakistan_friendly_only=pakistan_friendly_only,
        salary_only=salary_only,
        visa_support_only=visa_support_only,
        custom_keywords=custom_keywords,
    )
    client_access_links = build_dashboard_client_links(
        profile=selected,
        search_mode=search_mode,
        offer_type=offer_type,
        client_type=client_type,
        counterparty_type=counterparty_type,
        region_preference=region_preference,
        contact_goal=contact_goal,
        posted_within=posted_within,
        verified_payment_only=verified_payment_only,
        trust_signal=trust_signal,
        company_size=company_size,
        custom_keywords=custom_keywords,
    )
    outreach_links = build_outreach_links(selected, offer_type)
    active_filter_badges = build_active_filter_badges(
        search_mode=search_mode,
        offer_type=offer_type,
        client_type=client_type,
        contact_goal=contact_goal,
        posted_within=posted_within,
        counterparty_type=counterparty_type,
        trust_signal=trust_signal,
        company_size=company_size,
        proposal_pressure=proposal_pressure,
        verified_payment_only=verified_payment_only,
    )

    jobs = fetch_jobs_from_sources(selected_sources, limit_per_source=80) if not external_only_mode and not client_only_mode else []
    candidate_text = " ".join([selected.skills_csv, selected.cv_text])
    matches = rank_jobs(candidate_text=candidate_text, jobs=jobs, top_k=40)
    enriched_matches = []
    for item in matches:
        job = item["job"]
        level = score_to_level(float(item["score"]))
        if backend_only and not is_backend_job(job):
            continue
        if remote_only and not is_remote_job(job):
            continue
        if junior_only and not is_junior_friendly_job(job):
            continue
        if pakistan_friendly_only and not is_pakistan_friendly_job(job):
            continue
        if salary_only and not has_salary_signal(job):
            continue
        if visa_support_only and not has_visa_signal(job):
            continue
        if not matches_posted_window(job, posted_within):
            continue
        if not matches_region(job, region_preference):
            continue
        if not matches_search_focus(job, search_focus):
            continue
        if min_match_level in {"high", "medium", "low"}:
            if min_match_level == "high" and level != "high":
                continue
            if min_match_level == "medium" and level not in {"high", "medium"}:
                continue
        enriched_matches.append(
            {
                **item,
                "assisted_links": build_assisted_links(
                    job["title"],
                    job["company"],
                    candidate_name=selected.full_name,
                    candidate_email=selected.email,
                    resume_url=selected.resume_url,
                    portfolio_url=selected.portfolio_url,
                    github_username=selected.github_username,
                ),
                "match_level": level,
                "fit_score": compute_fit_score(float(item["score"])),
                "trust_score": compute_trust_score(job),
                "why_matched": build_match_reasons(job, item.get("matched_skills", []), item.get("missing_skills", [])),
                "is_remote": is_remote_job(job),
                "is_junior_friendly": is_junior_friendly_job(job),
                "is_pakistan_friendly": is_pakistan_friendly_job(job),
                "has_salary": has_salary_signal(job),
                "has_visa_support": has_visa_signal(job),
                "search_query": f"{job['title']} {job['company']}",
            }
        )
    enriched_matches = enriched_matches[:top_k]
    match_summary = build_match_summary(enriched_matches)
    client_search_results = []
    if active_ui_mode == "sell":
        client_search_results = build_client_search_results(
            live_search_links=live_search_links,
            client_access_links=client_access_links,
            offer_type=offer_type,
            client_type=client_type,
            contact_goal=contact_goal,
            counterparty_type=counterparty_type,
            custom_keywords=custom_keywords,
            top_k=top_k,
        )

    error = ""
    usage = None
    try:
        ensure_can_consume_matches(session, selected, len(enriched_matches))
        usage = consume_matches(session, selected, len(enriched_matches))
    except ValueError as exc:
        error = str(exc)
        enriched_matches = []

    return templates.TemplateResponse(
        "dashboard.html",
        {
            "request": request,
            "profiles": profiles,
            "selected": selected,
            "matches": enriched_matches,
            "usage": usage,
            "error": error,
            "sources": ",".join(selected_sources),
            "applications": build_tracker_entries(applications),
            "client_leads": build_client_lead_entries(client_leads),
            "saved_filter_presets": build_filter_preset_entries(filter_presets),
            "status_summary": status_summary,
            "client_status_summary": client_status_summary,
            "company_summary": build_company_summary(applications),
            "follow_up_summary": build_follow_up_summary(applications),
            "client_follow_up_summary": build_client_follow_up_summary(client_leads),
            "live_search_links": live_search_links,
            "client_access_links": client_access_links,
            "outreach_links": outreach_links,
            "resume_library": resume_library,
            "platform_targets": platform_targets,
            "selected_platform_targets": selected_platform_targets,
            "external_only_mode": external_only_mode,
            "client_only_mode": client_only_mode,
            "active_ui_mode": active_ui_mode,
            "search_mode": search_mode,
            "offer_type": offer_type,
            "client_type": client_type,
            "demand_level": demand_level,
            "contact_goal": contact_goal,
                "counterparty_type": counterparty_type,
                "posted_within": posted_within,
                "verified_payment_only": verified_payment_only,
                "trust_signal": trust_signal,
                "company_size": company_size,
                "proposal_pressure": proposal_pressure,
            "custom_keywords": custom_keywords,
            "min_match_level": min_match_level,
            "backend_only": backend_only,
            "remote_only": remote_only,
            "junior_only": junior_only,
            "search_focus": search_focus,
            "pakistan_friendly_only": pakistan_friendly_only,
            "salary_only": salary_only,
            "visa_support_only": visa_support_only,
            "region_preference": region_preference,
            "match_summary": match_summary,
            "active_filter_badges": active_filter_badges,
            "client_search_results": client_search_results,
        },
    )


# ──────────────────────────────────────────────────────────────────
#  NEW FEATURE 1: Duplicate Check
#  URL: GET /dashboard/client-leads/check-duplicate?email=...&phone=...
# ──────────────────────────────────────────────────────────────────
@router.get("/dashboard/client-leads/check-duplicate")
def check_lead_duplicate(
    session: Annotated[Session, Depends(get_session)],
    email: Optional[str] = None,
    phone: Optional[str] = None,
):
    if email:
        row = session.exec(
            select(ClientLead).where(ClientLead.contact_email == email)
        ).first()
        if row:
            return {"duplicate": True, "company": row.company, "lead_name": row.lead_name}

    if phone:
        clean_input = phone.replace(" ", "").replace("-", "").replace("+", "")
        all_leads = session.exec(select(ClientLead)).all()
        for row in all_leads:
            if row.contact_phone:
                stored = row.contact_phone.replace(" ", "").replace("-", "").replace("+", "")
                if stored == clean_input:
                    return {"duplicate": True, "company": row.company, "lead_name": row.lead_name}

    return {"duplicate": False}


# ──────────────────────────────────────────────────────────────────
#  NEW FEATURE 2: CSV Export
#  URL: GET /dashboard/client-leads/export-csv?candidate_id=1
# ──────────────────────────────────────────────────────────────────
@router.get("/dashboard/client-leads/export-csv")
def export_leads_csv(
    candidate_id: int,
    session: Annotated[Session, Depends(get_session)],
):
    leads = session.exec(
        select(ClientLead)
        .where(ClientLead.candidate_id == candidate_id)
        .order_by(ClientLead.updated_at.desc())
    ).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "ID", "Company", "Contact Person", "Source",
        "Email", "WhatsApp/Phone", "Channel",
        "Offer Type", "Status", "Follow-up Date",
        "Notes", "Lead URL", "Updated At",
    ])
    for lead in leads:
        writer.writerow([
            lead.id,
            lead.company,
            lead.lead_name or "",
            lead.source or "",
            lead.contact_email or "",
            lead.contact_phone or "",
            lead.contact_channel or "",
            lead.offer_type or "",
            lead.status,
            lead.follow_up_date or "",
            lead.notes or "",
            lead.lead_url or "",
            lead.updated_at.strftime("%Y-%m-%d %H:%M") if lead.updated_at else "",
        ])

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=leads.csv"},
    )




# ══════════════════════════════════════════════════════════════════
#  LEAD HUNTER PRO — Maximum Power Scraper
#  Sources: GitHub (500+), Dev.to, HackerNews, IndieHackers,
#           RSS Feeds, Remotive, WeWorkRemotely, Freelancer boards
# ══════════════════════════════════════════════════════════════════

import requests as _req
from bs4 import BeautifulSoup
import xml.etree.ElementTree as ET

_EMAIL_RE = re.compile(r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,7}')
# Strict E.164: leading +, first digit 1-9 (no leading zero), 7-15 digits
# total -- used to gate the "WhatsApp only" export down to numbers actually
# safe to hand to wa.me/, not whatever _WA_RE's looser text-pattern regexes
# above let through into the registry.
_E164_RE = re.compile(r'^\+[1-9]\d{6,14}$')
_WA_RE    = re.compile(
    r'wa\.me/(\d{7,15})'
    r'|whatsapp\.com/send\?phone=(\d{7,15})'
    r'|whatsapp[:\s\-]+(\+?[\d][\d\s\-]{9,14})'
    r'|\bwa[:\s\-]+(\+?[\d][\d\s\-]{9,14})'
    r'|\+92[\s\-]?3\d{2}[\s\-]?\d{7}'
    r'|\+1[\s\-]?\(?\d{3}\)?[\s\-]?\d{3}[\s\-]?\d{4}'
    r'|\+44[\s\-]?\d{10}'
    r'|\+91[\s\-]?\d{10}'
    r'|\+971[\s\-]?\d{9}',
    re.IGNORECASE
)

_SKIP_DOMAINS = {
    'example.com','test.com','email.com','domain.com','sentry.io',
    'github.com','google.com','cloudflare.com','w3.org','schema.org',
    'npmjs.com','pypi.org','jquery.com','wordpress.org','wix.com',
    'shopify.com','amazonaws.com','vercel.app','netlify.app',
    'heroku.com','railway.app','render.com','squarespace.com',
    'noreply.com','no-reply.com','notifications.com','support.com',
    'mailer.com','sendgrid.net','mailchimp.com','mailgun.org',
    'yourdomain.com','yoursite.com','mydomain.com','mycompany.com',
    'yourcompany.com','acme.com','foo.com','bar.com','company.com',
}

# RFC 2606 reserved TLDs — always example/placeholder, never real registrable mail.
# Catches tutorial/doc addresses like "alice@a.test" regardless of the domain name.
_RESERVED_TLDS = {'test', 'example', 'invalid', 'localhost'}

_HDRS = [
    {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'},
    {'User-Agent':'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36'},
    {'User-Agent':'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36'},
]

def _hdr(): return random.choice(_HDRS)

def _gh_headers() -> dict:
    """GitHub API headers — adds auth when GITHUB_TOKEN is configured, which
    lifts the unauthenticated rate limit (10/min search, 60/hr core) to
    30/min and 5000/hr and avoids silent 403s during a hunt."""
    headers = {'Accept': 'application/vnd.github.v3+json'}
    if settings.github_token:
        headers['Authorization'] = f'Bearer {settings.github_token}'
    return headers

def _mask_github_token(token: str) -> str:
    if len(token) < 8:
        return ""
    return f"{token[:4]}…{token[-4:]}"

def _gh_rate_limit_message() -> str:
    """GitHub scrapers used to swallow 403s and return an empty lead list,
    which looked exactly like a normal 'no matches this batch' result — the
    hunt appeared to run fine while silently finding nothing. Raise this
    instead so it surfaces as a real error in the batch result."""
    if settings.github_token:
        return "GitHub API rate limit reached even with a token configured — try again in a few minutes."
    return (
        "GitHub API rate limit reached (60 requests/hour without a token). "
        "Add a GITHUB_TOKEN in Settings to raise this to 5000/hour."
    )

# GitHub's Search API (/search/users, /search/repositories) has its own
# 30 requests/minute limit (10/min unauthenticated) that a GITHUB_TOKEN does
# NOT raise — it only raises the separate 60/hr -> 5,000/hr *core* limit
# (per-profile /users/{login} lookups, README fetches, etc). With several
# concurrent hunt lanes (see dashboard.html's LHP_CONCURRENCY / auto_hunt.py's
# PLATFORMS_PER_TICK) each capable of hitting a search endpoint at once, 30/min
# gets exceeded fast even with a token configured and every individual call
# legitimate — confirmed against a real hunt run repeatedly rate-limited
# despite a token being set. Shared across every thread hitting either search
# endpoint, not per-source, since it's GitHub's own limit that's shared too.
_GH_SEARCH_LOCK = threading.Lock()
_GH_SEARCH_TIMESTAMPS: deque = deque()
_GH_SEARCH_MAX_PER_MINUTE = 25


def _throttle_github_search() -> None:
    """Blocks the calling thread until a search-API slot is free, rather
    than firing anyway and eating a 403 — a hunt paces itself instead of
    repeatedly tripping the limit. Call this immediately before every
    request to api.github.com/search/*."""
    while True:
        with _GH_SEARCH_LOCK:
            now = time.monotonic()
            while _GH_SEARCH_TIMESTAMPS and now - _GH_SEARCH_TIMESTAMPS[0] > 60:
                _GH_SEARCH_TIMESTAMPS.popleft()
            if len(_GH_SEARCH_TIMESTAMPS) < _GH_SEARCH_MAX_PER_MINUTE:
                _GH_SEARCH_TIMESTAMPS.append(now)
                return
            wait_s = 60 - (now - _GH_SEARCH_TIMESTAMPS[0]) + 0.05
        time.sleep(max(wait_s, 0.05))


def _valid_email(e: str) -> bool:
    d = e.split('@')[-1].lower()
    tld = d.rsplit('.', 1)[-1]
    return (
        d not in _SKIP_DOMAINS
        and tld not in _RESERVED_TLDS
        and '..' not in e
        and len(e) < 80
        and '.' in d
        and not e.startswith('.')
        and not e.endswith('.')
    )

def _extract(text: str) -> tuple[list, list]:
    """Extract emails and WhatsApp numbers from text"""
    emails = list({e for e in _EMAIL_RE.findall(text) if _valid_email(e)})

    wa = []
    # wa.me links
    for m in re.finditer(r'wa\.me/(\d{7,15})', text, re.I):
        wa.append('+' + m.group(1))
    # whatsapp.com/send?phone=
    for m in re.finditer(r'whatsapp\.com/send\?phone=(\d{7,15})', text, re.I):
        wa.append('+' + m.group(1))
    # "whatsapp: +923001234567" or "wa: +923001234567"
    for m in re.finditer(r'(?:whatsapp|wa)[:\s\-]+(\+?[\d][\d\s\-]{9,14})', text, re.I):
        n = re.sub(r'[\s\-]', '', m.group(1))
        if len(n) >= 10:
            wa.append(n if n.startswith('+') else '+' + n)
    # Phone numbers with country codes
    for pattern in [
        r'\+92[\s\-]?3\d{2}[\s\-]?\d{7}',  # Pakistan
        r'\+1[\s\-]?\(?\d{3}\)?[\s\-]?\d{3}[\s\-]?\d{4}',  # USA
        r'\+44[\s\-]?\d{10}',  # UK
        r'\+91[\s\-]?\d{10}',  # India
        r'\+971[\s\-]?\d{9}',  # UAE
        r'\+966[\s\-]?\d{9}',  # Saudi
        r'\+61[\s\-]?\d{9}',   # Australia
        r'\+49[\s\-]?\d{10,11}', # Germany
    ]:
        for m in re.finditer(pattern, text):
            n = re.sub(r'[\s\-\(\)]', '', m.group(0))
            if len(n) >= 10:
                wa.append(n)

    return emails, list(set(wa))

def _desig(bio: str) -> str:
    if not bio: return ''
    for pat in [
        r'([\w\s]+developer)', r'([\w\s]+designer)',
        r'(freelance[\w\s]+)', r'([\w\s]+engineer)',
        r'([\w\s]+marketer)', r'([\w\s]+creator)',
        r'([\w\s]+consultant)', r'([\w\s]+writer)',
        r'([\w\s]+coach)', r'([\w\s]+seller)',
        r'([\w\s]+founder)', r'([\w\s]+maker)',
    ]:
        m = re.search(pat, bio, re.IGNORECASE)
        if m: return m.group(0).strip()[:50]
    return bio.split('.')[0][:50]


# ── SOURCE 1: GitHub Profiles (best for emails) ──────────────────

GH_SEARCH_QUERIES = [
    # Freelancers with emails
    'freelance developer email hire',
    'hire me developer email contact',
    'open to work email developer',
    'available for hire email developer',
    'freelance designer email hire',
    'freelance marketer email contact',
    'indie maker email contact',
    'solopreneur email hire',
    'python developer freelance email',
    'react developer freelance email',
    'fullstack developer freelance email',
    'web developer hire email contact',
    'mobile developer freelance email',
    'wordpress developer email hire',
    'shopify developer email contact',
    'django developer freelance email',
    'node developer freelance email',
    'flutter developer freelance email',
    'data scientist freelance email',
    'machine learning engineer email',
    'devops engineer freelance email',
    'blockchain developer email hire',
    'graphic designer email whatsapp hire',
    'UI UX designer email contact hire',
    'content creator email whatsapp',
    'digital marketer email whatsapp hire',
    'SEO expert email hire contact',
    'video editor email hire whatsapp',
    'copywriter email whatsapp hire',
    'virtual assistant email hire',
    'translator freelance email contact',
    'data entry freelance email whatsapp',
    'bookkeeper accountant freelance email',
    'voice over artist email whatsapp hire',
    '3d animator freelance email contact',
    'game developer freelance email hire',
    'QA tester freelance email contact',
    'technical writer freelance email',
    'social media manager email whatsapp',
    'ecommerce consultant email hire',
    'cybersecurity consultant freelance email',
    'aws cloud consultant email hire',
    'no code developer bubble webflow email',
    'photographer videographer freelance email',
    'legal consultant freelance email contact',
    'recruiter hr consultant freelance email',
    'business analyst freelance email hire',
    'product manager freelance email contact',
]

def _fetch_github_profiles(keyword: str, page: int = 1) -> list[dict]:
    """GitHub user profiles with public emails — page support for 500+ results"""
    leads = []
    try:
        _throttle_github_search()
        url = f"https://api.github.com/search/users?q={_req.utils.quote(keyword)}&per_page=30&page={page}"
        r = _req.get(url, headers=_gh_headers(), timeout=15)
    except Exception:
        logger.exception("github profiles: search request failed for %r", keyword)
        return leads

    if r.status_code == 422: return leads  # Bad query
    if r.status_code == 403:
        raise RuntimeError(_gh_rate_limit_message())
    if r.status_code != 200:
        logger.warning("github: search %r returned %s", keyword, r.status_code)
        return leads

    users = r.json().get('items', [])
    profile_calls = 0
    rate_limited_calls = 0
    for u in users[:20]:
        try:
            pr = _req.get(
                f"https://api.github.com/users/{u['login']}",
                headers=_gh_headers(),
                timeout=10
            )
        except Exception:
            continue
        profile_calls += 1
        if pr.status_code == 403:
            rate_limited_calls += 1
            continue
        if pr.status_code != 200: continue
        try:
            p = pr.json()

            # Get all possible contact info
            email = p.get('email') or ''
            bio = f"{p.get('bio') or ''} {p.get('blog') or ''} {p.get('location') or ''} {p.get('company') or ''}"
            bio_emails, bio_wa = _extract(bio)

            if not email and bio_emails:
                email = bio_emails[0]
            wa = bio_wa[0] if bio_wa else ''

            # Check blog URL for email
            blog = p.get('blog') or ''
            if blog and not email:
                be, bw = _extract(blog)
                email = be[0] if be else ''
                if not wa and bw: wa = bw[0]

            # GitHub's profile "website" field (blog) is usually the
            # candidate's own personal/business site — a real crawl target
            # for Pass 2's wa.me/mailto discovery (see website_crawler.py),
            # not just more text to regex-scan. Only worth passing on when
            # it's a real absolute URL, not a bare handle some profiles put
            # here instead (e.g. "@someuser").
            website = blog if blog.startswith(('http://', 'https://')) else ''

            if email or wa or website:
                leads.append({
                    'name': p.get('name') or p.get('login', ''),
                    'designation': _desig(p.get('bio') or 'Developer'),
                    'email': email,
                    'whatsapp': wa,
                    'source': 'github',
                    'url': p.get('html_url', ''),
                    'website': website,
                    'notes': (p.get('bio') or '')[:100],
                    'location': p.get('location') or '',
                })
            time.sleep(0.2)
        except Exception:
            continue

    # A search that returned real candidates but got 403'd on every single
    # profile lookup is the exhausted-quota case, not "nobody had contact
    # info" — surface it instead of returning an indistinguishable empty list.
    if profile_calls and rate_limited_calls == profile_calls:
        raise RuntimeError(_gh_rate_limit_message())
    return leads


def _fetch_github_readme(keyword: str, page: int = 1) -> list[dict]:
    """Search GitHub READMEs for contact info"""
    leads = []
    try:
        _throttle_github_search()
        q = f"{keyword} email in:readme"
        url = f"https://api.github.com/search/repositories?q={_req.utils.quote(q)}&per_page=20&sort=updated&page={page}"
        r = _req.get(url, headers=_gh_headers(), timeout=15)
        if r.status_code == 403:
            raise RuntimeError(_gh_rate_limit_message())
        if r.status_code != 200:
            logger.warning("github readme: search %r returned %s", keyword, r.status_code)
            return leads

        for repo in r.json().get('items', [])[:12]:
            try:
                owner = repo.get('owner', {}).get('login', '')
                rname = repo.get('name', '')
                # Try main then master branch
                for branch in ['main', 'master']:
                    rm = _req.get(
                        f"https://raw.githubusercontent.com/{owner}/{rname}/{branch}/README.md",
                        timeout=8
                    )
                    if rm.status_code == 200:
                        emails, wa = _extract(rm.text)
                        if emails or wa:
                            leads.append({
                                'name': owner,
                                'designation': _desig(repo.get('description') or 'Developer'),
                                'email': emails[0] if emails else '',
                                'whatsapp': wa[0] if wa else '',
                                'source': 'github',
                                'url': repo.get('html_url', ''),
                                'notes': (repo.get('description') or '')[:100],
                            })
                        break
                time.sleep(0.25)
            except Exception:
                continue
    except RuntimeError:
        raise
    except Exception:
        logger.exception("github readme: fetch failed for %r", keyword)
    return leads


# ── SOURCE 2: Dev.to (great for developer emails) ────────────────

DEVTO_TAGS = [
    'forhire', 'freelance', 'hiring', 'opentowork', 'career',
    'developer', 'webdev', 'python', 'javascript', 'react',
    'node', 'fullstack', 'backend', 'frontend', 'devops',
    'design', 'ux', 'marketing', 'productivity', 'startup',
]

def _fetch_devto_articles(tag: str, page: int = 1) -> list[dict]:
    """Dev.to articles — fetch full body for contact info"""
    leads = []
    try:
        r = _req.get(
            f"https://dev.to/api/articles?tag={tag}&per_page=20&page={page}",
            headers=_hdr(), timeout=12
        )
        if r.status_code != 200: return leads
        for a in r.json():
            try:
                # Get full article with body
                art_id = a.get('id', '')
                if not art_id: continue
                ar = _req.get(
                    f"https://dev.to/api/articles/{art_id}",
                    headers=_hdr(), timeout=10
                )
                if ar.status_code != 200: continue
                full = ar.json()
                body = full.get('body_markdown', '') or full.get('body_html', '') or ''
                # Clean HTML tags
                body_clean = re.sub(r'<[^>]+>', ' ', body)
                txt = f"{a.get('title','')} {a.get('description','')} {body_clean}"
                emails, wa = _extract(txt)
                if emails or wa:
                    user = a.get('user', {})
                    leads.append({
                        'name': user.get('name', ''),
                        'designation': _desig(a.get('title', '')),
                        'email': emails[0] if emails else '',
                        'whatsapp': wa[0] if wa else '',
                        'source': 'devto',
                        'url': a.get('url', ''),
                        'notes': a.get('title', '')[:100],
                    })
                time.sleep(0.15)
            except Exception:
                continue
    except Exception:
        logger.exception("devto: fetch failed for tag %r", tag)
    return leads


# ── SOURCE 3: HackerNews (great for dev emails) ──────────────────

def _fetch_hackernews() -> list[dict]:
    """HackerNews Who Is Hiring + Freelancer threads"""
    leads = []
    # Fetching comments is one sequential HTTP call per comment — cap both the
    # per-story comment count and total wall-clock time so a busy/slow thread
    # (which can have hundreds of kids) can never turn one batch into minutes
    # of blocking requests.
    deadline = time.time() + 18
    try:
        # Search for hiring/freelance posts via Algolia
        for query in ['Ask HN Who is hiring', 'Ask HN Freelancer']:
            if time.time() > deadline: break
            r = _req.get(
                f'https://hn.algolia.com/api/v1/search?query={_req.utils.quote(query)}&tags=story&hitsPerPage=3',
                timeout=12
            )
            if r.status_code != 200:
                logger.warning("hackernews: search for %r returned %s", query, r.status_code)
                continue
            for hit in r.json().get('hits', [])[:2]:
                if time.time() > deadline: break
                story_id = hit.get('objectID', '')
                if not story_id: continue
                # Get comments
                sr = _req.get(
                    f'https://hacker-news.firebaseio.com/v0/item/{story_id}.json',
                    timeout=10
                )
                if sr.status_code != 200: continue
                kids = (sr.json().get('kids') or [])[:15]
                for kid_id in kids:
                    if time.time() > deadline: break
                    try:
                        kr = _req.get(
                            f'https://hacker-news.firebaseio.com/v0/item/{kid_id}.json',
                            timeout=6
                        )
                        if kr.status_code != 200: continue
                        k = kr.json()
                        txt = re.sub(r'<[^>]+>', ' ', k.get('text', '') or '')
                        emails, wa = _extract(txt)
                        if emails or wa:
                            leads.append({
                                'name': k.get('by', ''),
                                'designation': _desig(txt[:200]),
                                'email': emails[0] if emails else '',
                                'whatsapp': wa[0] if wa else '',
                                'source': 'hackernews',
                                'url': f"https://news.ycombinator.com/item?id={k.get('id','')}",
                                'notes': txt[:100],
                            })
                        time.sleep(0.05)
                    except Exception:
                        continue
    except Exception:
        logger.exception("hackernews: fetch failed")
    return leads


# ── SOURCE 4: IndieHackers ───────────────────────────────────────

def _fetch_indiehackers() -> list[dict]:
    leads = []
    try:
        r = _req.get('https://www.indiehackers.com/api/posts?limit=50', timeout=12)
        if r.status_code != 200: return leads
        for p in r.json()[:30]:
            txt = f"{p.get('title','')} {p.get('content','')}"
            emails, wa = _extract(txt)
            if emails or wa:
                leads.append({
                    'name': p.get('userDisplayName', ''),
                    'designation': _desig(p.get('title', '')),
                    'email': emails[0] if emails else '',
                    'whatsapp': wa[0] if wa else '',
                    'source': 'indiehackers',
                    'url': f"https://www.indiehackers.com{p.get('url','')}",
                    'notes': p.get('title', '')[:100],
                })
    except Exception:
        logger.exception("indiehackers: fetch failed")
    return leads


# ── SOURCE 5: RSS Feeds from Job Boards ─────────────────────────

def _fetch_remotive_rss() -> list[dict]:
    """Remotive.com RSS — remote jobs with company emails"""
    leads = []
    try:
        r = _req.get('https://remotive.com/remote-jobs/feed', headers=_hdr(), timeout=15)
        if r.status_code != 200: return leads
        root = ET.fromstring(r.content)
        ns = {'content': 'http://purl.org/rss/1.0/modules/content/'}
        for item in root.findall('.//item')[:30]:
            title = (item.findtext('title') or '')
            company = (item.findtext('author') or '')
            link = (item.findtext('link') or '')
            desc = (item.findtext('description') or '')
            content = (item.find('content:encoded', ns) or item.find('{http://purl.org/rss/1.0/modules/content/}encoded'))
            body = content.text if content is not None else ''
            txt = f"{title} {company} {desc} {body}"
            txt_clean = re.sub(r'<[^>]+>', ' ', txt)
            emails, wa = _extract(txt_clean)
            if emails or wa:
                leads.append({
                    'name': company or title[:40],
                    'designation': _desig(title),
                    'email': emails[0] if emails else '',
                    'whatsapp': wa[0] if wa else '',
                    'source': 'remotive',
                    'url': link,
                    'notes': title[:100],
                })
    except Exception:
        logger.exception("remotive: fetch failed")
    return leads


def _fetch_weworkremotely_rss() -> list[dict]:
    """WeWorkRemotely RSS feeds"""
    leads = []
    feeds = [
        'https://weworkremotely.com/remote-jobs.rss',
        'https://weworkremotely.com/categories/remote-programming-jobs.rss',
        'https://weworkremotely.com/categories/remote-design-jobs.rss',
    ]
    for feed_url in feeds:
        try:
            r = _req.get(feed_url, headers=_hdr(), timeout=12)
            if r.status_code != 200: continue
            root = ET.fromstring(r.content)
            for item in root.findall('.//item')[:20]:
                title = (item.findtext('title') or '')
                link = (item.findtext('link') or '')
                desc = item.findtext('{http://purl.org/rss/1.0/modules/content/}encoded') or item.findtext('description') or ''
                txt_clean = re.sub(r'<[^>]+>', ' ', f"{title} {desc}")
                emails, wa = _extract(txt_clean)
                if emails or wa:
                    leads.append({
                        'name': title.split('at ')[-1][:50] if ' at ' in title else title[:50],
                        'designation': _desig(title),
                        'email': emails[0] if emails else '',
                        'whatsapp': wa[0] if wa else '',
                        'source': 'weworkremotely',
                        'url': link,
                        'notes': title[:100],
                    })
        except Exception:
            logger.exception("weworkremotely: fetch failed for %s", feed_url)
            continue
    return leads


def _fetch_github_jobs_rss() -> list[dict]:
    """GitHub discussions and community posts"""
    leads = []
    try:
        # GitHub explore feed
        _throttle_github_search()
        r = _req.get(
            'https://api.github.com/search/issues?q=freelance+email+hire+label:hiring&sort=created&order=desc&per_page=30',
            headers=_gh_headers(),
            timeout=12
        )
        if r.status_code != 200:
            logger.warning("github issues: search returned %s", r.status_code)
            return leads
        for issue in r.json().get('items', [])[:20]:
            txt = f"{issue.get('title','')} {issue.get('body','') or ''}"
            txt_clean = re.sub(r'<[^>]+>', ' ', txt)
            emails, wa = _extract(txt_clean)
            if emails or wa:
                leads.append({
                    'name': issue.get('user',{}).get('login',''),
                    'designation': _desig(issue.get('title','')),
                    'email': emails[0] if emails else '',
                    'whatsapp': wa[0] if wa else '',
                    'source': 'github',
                    'url': issue.get('html_url',''),
                    'notes': issue.get('title','')[:100],
                })
    except Exception:
        logger.exception("github issues: fetch failed")
    return leads


def _fetch_reddit_rss(subreddit: str) -> list[dict]:
    """Reddit via RSS — different from JSON API, often works"""
    leads = []
    try:
        url = f"https://www.reddit.com/r/{subreddit}/new/.rss?limit=50"
        hdrs = {
            'User-Agent': 'Mozilla/5.0 (compatible; RSS reader)',
            'Accept': 'application/rss+xml, application/xml, text/xml',
        }
        r = _req.get(url, headers=hdrs, timeout=15)
        if r.status_code != 200:
            logger.warning("reddit: r/%s returned %s", subreddit, r.status_code)
            return leads
        root = ET.fromstring(r.content)
        ns = {
            'atom': 'http://www.w3.org/2005/Atom',
            'content': 'http://purl.org/rss/1.0/modules/content/',
        }
        entries = root.findall('.//entry') or root.findall('.//item')
        for entry in entries[:30]:
            title = (
                (entry.find('atom:title', ns) or entry.find('title'))
            )
            title_txt = title.text if title is not None else ''
            content_el = (
                entry.find('content') or
                entry.find('atom:content', ns) or
                entry.find('{http://purl.org/rss/1.0/modules/content/}encoded')
            )
            body = content_el.text if content_el is not None else ''
            author_el = entry.find('atom:author/atom:name', ns) or entry.find('author')
            author = author_el.text if author_el is not None else ''
            link_el = entry.find('atom:link', ns) or entry.find('link')
            link = link_el.get('href','') if link_el is not None else ''
            txt_clean = re.sub(r'<[^>]+>', ' ', f"{title_txt} {body}")
            emails, wa = _extract(txt_clean)
            if emails or wa:
                leads.append({
                    'name': author,
                    'designation': _desig(title_txt),
                    'email': emails[0] if emails else '',
                    'whatsapp': wa[0] if wa else '',
                    'source': 'reddit',
                    'url': link,
                    'notes': title_txt[:100],
                })
    except Exception:
        logger.exception("reddit: fetch failed for r/%s", subreddit)
    return leads


# ── SOURCE 6: Public Portfolio Sites ────────────────────────────

def _fetch_carrd_profiles(keyword: str) -> list[dict]:
    """Carrd.co portfolio pages via Google search-like approach"""
    leads = []
    # These are already indexed public pages
    portfolio_urls = [
        f"https://dev.to/t/forhire",
        f"https://www.indiehackers.com/group/forhire",
    ]
    try:
        for url in portfolio_urls:
            r = _req.get(url, headers=_hdr(), timeout=12)
            if r.status_code != 200: continue
            soup = BeautifulSoup(r.text, 'html.parser')
            txt = soup.get_text(' ', strip=True)
            emails, wa = _extract(txt)
            for email in emails[:5]:
                leads.append({
                    'name': '',
                    'designation': 'Freelancer',
                    'email': email,
                    'whatsapp': wa[0] if wa else '',
                    'source': 'portfolio',
                    'url': url,
                    'notes': f'From {url}',
                })
    except Exception:
        pass
    return leads


# ── API Endpoints ─────────────────────────────────────────────────

GH_KEYWORDS = GH_SEARCH_QUERIES  # alias

@router.get("/api/scrape/test")
async def test_scraper():
    """Test which sources are accessible"""
    results = {}
    tests = [
        ('reddit_rss', 'https://www.reddit.com/r/forhire/new/.rss?limit=5', {'User-Agent':'Mozilla/5.0 (compatible; RSS)','Accept':'application/rss+xml'}),
        ('github', 'https://api.github.com/search/users?q=freelance+developer+email&per_page=3', {'Accept':'application/vnd.github.v3+json'}),
        ('devto', 'https://dev.to/api/articles?tag=forhire&per_page=3', {}),
        ('hackernews', 'https://hn.algolia.com/api/v1/search?query=freelancer&tags=story&hitsPerPage=3', {}),
        ('indiehackers', 'https://www.indiehackers.com/api/posts?limit=3', {}),
        ('remotive_rss', 'https://remotive.com/remote-jobs/feed', _hdr()),
        ('weworkremotely', 'https://weworkremotely.com/remote-jobs.rss', _hdr()),
    ]
    for name, url, hdrs in tests:
        try:
            r = _req.get(url, headers=hdrs, timeout=8)
            results[name] = {'status': r.status_code, 'ok': r.status_code == 200}
        except Exception as e:
            results[name] = {'status': 'error', 'error': str(e)[:60], 'ok': False}
    return JSONResponse(results)


@router.get("/api/scrape/platforms")
async def scrape_platforms():
    """212+ hunting platforms grouped by source type."""
    from app.services.lead_hunter_engine import list_platforms, summary

    return JSONResponse({"platforms": list_platforms(), "summary": summary()})


@router.get("/api/scrape/hunt-plan")
async def scrape_hunt_plan(chips: str = Query(default="github,reddit,devto,misc")):
    """Dynamic hunt plan for enabled UI chips."""
    from app.services.lead_hunter_engine import build_hunt_plan, summary

    enabled = [c.strip() for c in chips.split(",") if c.strip()]
    plan = build_hunt_plan(enabled)
    return JSONResponse({"plan": plan, "total_batches": sum(p["batches"] for p in plan), "summary": summary()})


@router.get("/api/scrape/registry")
async def scrape_registry(session: Annotated[Session, Depends(get_session)]):
    """Permanent hunted email/WhatsApp registry — never search these again."""
    from app.services.lead_hunter_registry import load_known_keys, registry_stats

    emails, whatsapps = load_known_keys(session)
    stats = registry_stats(session)
    return JSONResponse(
        {
            **stats,
            "emails": sorted(emails),
            "whatsapp": sorted(whatsapps),
        }
    )


@router.post("/api/scrape/session/start")
async def scrape_session_start(
    session: Annotated[Session, Depends(get_session)],
    target: str = Form(default=""),
    keywords: str = Form(default=""),
    location: str = Form(default=""),
    chips: str = Form(default=""),
):
    """Open a hunt session so every contact found during this run can be
    recovered and exported from disk later, even after a restart."""
    from app.services.lead_hunter_registry import start_hunt_session

    hunt_session = start_hunt_session(
        session, target=target, keywords=keywords, location=location, chips_csv=chips
    )
    return JSONResponse({"id": hunt_session.id, "started_at": hunt_session.started_at.isoformat()})


@router.post("/api/scrape/session/{session_id}/end")
async def scrape_session_end(
    session_id: int,
    session: Annotated[Session, Depends(get_session)],
    status: str = Form(default="completed"),
):
    from app.services.lead_hunter_registry import end_hunt_session

    hunt_session = end_hunt_session(session, session_id, status=status)
    if not hunt_session:
        raise HTTPException(status_code=404, detail="Unknown hunt session")
    return JSONResponse({"id": hunt_session.id, "status": hunt_session.status})


@router.get("/api/scrape/session/latest")
async def scrape_session_latest(session: Annotated[Session, Depends(get_session)]):
    """The most recent hunt session — lets the dashboard offer to recover an
    in-progress or just-finished hunt after a reload/restart."""
    from app.services.lead_hunter_registry import get_latest_session

    latest = get_latest_session(session)
    return JSONResponse(latest or {})


@router.get("/api/scrape/all-time/{kind}")
async def scrape_all_time_download(kind: str):
    """Downloads the cumulative all-time contact file — every unique
    email/WhatsApp ever hunted across every session, not just the current
    one. For the user's own reuse (e.g. importing into their own outreach
    tool); this is not a bulk export/resale feature."""
    from app.paths import data_dir

    filenames = {"emails": "emails_all_time.txt", "whatsapp": "whatsapp_all_time.txt"}
    filename = filenames.get(kind)
    if not filename:
        raise HTTPException(status_code=404, detail="Unknown all-time file")
    path = data_dir() / filename
    if not path.exists():
        return PlainTextResponse("", media_type="text/plain")
    return FileResponse(path, media_type="text/plain", filename=filename)


@router.get("/api/scrape/auto-hunt")
async def auto_hunt_status_api(session: Annotated[Session, Depends(get_session)]):
    from app.services.auto_hunt import get_auto_hunt_settings

    return JSONResponse(get_auto_hunt_settings(session))


@router.post("/api/scrape/auto-hunt")
async def auto_hunt_save_api(
    session: Annotated[Session, Depends(get_session)],
    enabled: str = Form(default=""),
    interval_hours: int = Form(default=4),
):
    from app.services.auto_hunt import set_auto_hunt_settings

    return JSONResponse(set_auto_hunt_settings(session, enabled == "1", interval_hours))


@router.get("/api/scrape/pool-health")
async def scrape_pool_health_api(session: Annotated[Session, Depends(get_session)]):
    """Total deduplicated pool size + growth + confidence breakdown — the
    metric that matters once hunting is a continuous background process
    rather than one-off sessions: a growing number, not a single big run."""
    from app.services.lead_hunter_registry import pool_health

    return JSONResponse(pool_health(session))


@router.get("/api/scrape/history")
async def scrape_history_api(session: Annotated[Session, Depends(get_session)]):
    """Every past hunt session with its contact counts — the Hunt History
    view. Re-export uses the existing /api/scrape/export?session_id=... and
    /api/scrape/session/{id}/leads routes, which already read straight from
    disk; nothing here re-runs a hunt."""
    from app.services.lead_hunter_registry import list_hunt_sessions

    return JSONResponse({"sessions": list_hunt_sessions(session)})


@router.get("/api/scrape/session/{session_id}/leads")
async def scrape_session_leads(session_id: int, session: Annotated[Session, Depends(get_session)]):
    """Full lead records for a hunt session, read straight from disk — the
    durable source of truth an export can always fall back to."""
    from app.services.lead_hunter_registry import get_session_leads

    return JSONResponse({"leads": get_session_leads(session, session_id)})


@router.get("/api/geo/search")
async def geo_search_api(q: str = Query(default="")):
    """Forward-geocode a place name (city/country) via Nominatim, for the location autocomplete."""
    query = q.strip()
    if not query:
        return JSONResponse({"results": []})

    import requests as _requests

    try:
        resp = _requests.get(
            "https://nominatim.openstreetmap.org/search",
            params={
                "q": query,
                "format": "jsonv2",
                "limit": 6,
                "addressdetails": 1,
                "accept-language": "en",
            },
            headers={"User-Agent": "JobMindMatch/1.0 (lead-hunter location picker)"},
            timeout=6,
        )
        resp.raise_for_status()
        items = resp.json()
    except Exception:
        return JSONResponse({"results": []})

    def _pick_city(address: dict) -> str:
        return address.get("city") or address.get("town") or address.get("village") or ""

    results = [
        {
            "display_name": item.get("display_name", ""),
            "lat": item.get("lat", ""),
            "lon": item.get("lon", ""),
            "country": (item.get("address") or {}).get("country", ""),
            "city": _pick_city(item.get("address") or {}),
        }
        for item in items
    ]
    return JSONResponse({"results": results})


@router.get("/api/geo/reverse")
async def geo_reverse_api(lat: float = Query(...), lon: float = Query(...)):
    """Reverse-geocode a map pin drop via Nominatim, for the location picker."""
    import requests as _requests

    try:
        resp = _requests.get(
            "https://nominatim.openstreetmap.org/reverse",
            params={
                "lat": lat,
                "lon": lon,
                "format": "jsonv2",
                "addressdetails": 1,
                "accept-language": "en",
            },
            headers={"User-Agent": "JobMindMatch/1.0 (lead-hunter location picker)"},
            timeout=6,
        )
        resp.raise_for_status()
        item = resp.json()
    except Exception:
        return JSONResponse({"display_name": "", "country": "", "city": ""})

    address = item.get("address") or {}
    return JSONResponse(
        {
            "display_name": item.get("display_name", ""),
            "country": address.get("country", ""),
            "city": address.get("city") or address.get("town") or address.get("village") or "",
        }
    )


@router.get("/api/geo/radius")
async def geo_radius_api(
    lat: float = Query(...),
    lon: float = Query(...),
    radius_km: float = Query(default=25, ge=1, le=300),
):
    """Real places (cities/towns/villages) inside a radius via Overpass — powers the
    Facebook-Ads-style radius picker. Distinct from Google Maps scraping: this reads
    OpenStreetMap's own public place index, not business listings."""
    import math
    import time

    import requests as _requests

    radius_m = int(radius_km * 1000)
    query = (
        "[out:json][timeout:9];"
        f'(node["place"~"^(city|town|village)$"](around:{radius_m},{lat},{lon}););'
        "out body 60;"
    )
    # Overpass is a free, shared community server — it can transiently rate-limit or
    # time out under load, so retry once before telling the user nothing was found.
    elements: list = []
    last_error = ""
    for attempt in range(2):
        try:
            resp = _requests.get(
                "https://overpass-api.de/api/interpreter",
                params={"data": query},
                headers={"User-Agent": "JobMindMatch/1.0 (lead-hunter radius picker)"},
                timeout=10,
            )
            resp.raise_for_status()
            elements = resp.json().get("elements", [])
            last_error = ""
            break
        except Exception as exc:
            last_error = str(exc)
            if attempt == 0:
                time.sleep(1)

    if last_error:
        return JSONResponse({"places": [], "error": "Location service is busy right now — try again in a moment."})

    def _distance_km(lat2: float, lon2: float) -> float:
        r = 6371.0
        dlat = math.radians(lat2 - lat)
        dlon = math.radians(lon2 - lon)
        a = (
            math.sin(dlat / 2) ** 2
            + math.cos(math.radians(lat)) * math.cos(math.radians(lat2)) * math.sin(dlon / 2) ** 2
        )
        return r * 2 * math.asin(math.sqrt(a))

    places = []
    seen_names = set()
    for el in elements:
        tags = el.get("tags", {})
        name = tags.get("name:en") or tags.get("name") or ""
        if not name or name in seen_names:
            continue
        seen_names.add(name)
        places.append(
            {
                "name": name,
                "lat": el.get("lat"),
                "lon": el.get("lon"),
                "distance_km": round(_distance_km(el.get("lat", lat), el.get("lon", lon)), 1),
            }
        )

    places.sort(key=lambda p: p["distance_km"])
    return JSONResponse({"places": places})


@router.post("/api/dnc/add")
async def dnc_add_api(
    session: Annotated[Session, Depends(get_session)],
    email: str = Form(default=""),
    whatsapp: str = Form(default=""),
    reason: str = Form(default=""),
):
    """Add an email/WhatsApp to the do-not-contact suppression list."""
    from app.models import DoNotContactEntry

    email_norm = email.strip().lower()
    whatsapp_norm = whatsapp.strip()
    if not email_norm and not whatsapp_norm:
        raise HTTPException(status_code=400, detail="email or whatsapp required")

    entry = DoNotContactEntry(email=email_norm, whatsapp=whatsapp_norm, reason=reason.strip())
    session.add(entry)
    session.commit()
    return JSONResponse({"ok": True})


@router.get("/api/dnc/list")
async def dnc_list_api(session: Annotated[Session, Depends(get_session)]):
    """List suppressed emails/WhatsApp numbers."""
    from app.models import DoNotContactEntry

    entries = session.exec(
        select(DoNotContactEntry).order_by(DoNotContactEntry.created_at.desc())
    ).all()
    return JSONResponse(
        {
            "entries": [
                {
                    "id": e.id,
                    "email": e.email,
                    "whatsapp": e.whatsapp,
                    "reason": e.reason,
                    "created_at": e.created_at.isoformat(),
                }
                for e in entries
            ]
        }
    )


@router.get("/api/scrape/leads")
async def scrape_leads(
    session: Annotated[Session, Depends(get_session)],
    target: str = Query(default="all"),
    keywords: str = Query(default=""),
    country: str = Query(default=""),
    location: str = Query(default=""),
    locations: str = Query(default=""),
    batch_size: int = Query(default=50),
    source: str = Query(default="github"),
    offset: int = Query(default=0),
    session_id: Optional[int] = Query(default=None),
):
    """Server-side lead scraper — filters against permanent hunted registry."""
    from app.services.lead_hunter_engine import run_scrape_batch

    kw = keywords.strip()
    loc = location.strip() or country.strip()
    loc_list = [p.strip() for p in locations.split(",") if p.strip()]

    try:
        # run_scrape_batch is a synchronous function that itself blocks on a
        # ThreadPoolExecutor future (up to _SOURCE_TIMEOUT_SECONDS) — calling
        # it directly here would freeze the whole asyncio event loop for
        # that long, stalling every other request the app is serving during
        # a hunt. Offload it to a worker thread instead (the sqlite
        # connection was opened with check_same_thread=False for exactly
        # this reason).
        result = await asyncio.to_thread(
            run_scrape_batch, session, source, offset, kw, loc, loc_list, hunt_session_id=session_id
        )
        return JSONResponse(result)
    except Exception as e:
        logger.exception("lead hunt: /api/scrape/leads failed for source=%s offset=%s", source, offset)
        return JSONResponse(
            {
                "leads": [],
                "count": 0,
                "source": source,
                "offset": offset,
                "skipped_known": 0,
                "registry_total": 0,
                "error": str(e),
            }
        )


@router.get("/api/license/status")
async def license_status_api(session: Annotated[Session, Depends(get_session)]):
    from app.services.license_service import license_status

    return JSONResponse(license_status(session))


@router.post("/api/license/activate")
async def license_activate_api(
    session: Annotated[Session, Depends(get_session)],
    activation_code: str = Form(default=""),
):
    from app.services.license_service import activate_license

    return JSONResponse(activate_license(session, activation_code))


@router.get("/api/settings/github-token")
async def github_token_status_api(session: Annotated[Session, Depends(get_session)]):
    row = session.get(AppSetting, "github_token")
    token = row.value if row else ""
    return JSONResponse({"configured": bool(token), "masked": _mask_github_token(token)})


@router.post("/api/settings/github-token")
async def github_token_save_api(
    session: Annotated[Session, Depends(get_session)],
    token: str = Form(default=""),
):
    # A blank submit clears the saved token, falling back to unauthenticated
    # GitHub calls (60/hr) rather than being treated as "no change" — that's
    # the only way to remove a token once one is saved.
    cleaned = token.strip()
    session.merge(AppSetting(key="github_token", value=cleaned))
    session.commit()
    # settings is a shared singleton imported by _gh_headers(), github_client.py,
    # and _resolve_latest_release() below — mutating it here makes the new
    # token take effect immediately, no restart needed.
    settings.github_token = cleaned
    return JSONResponse({"ok": True, "configured": bool(cleaned), "masked": _mask_github_token(cleaned)})


@router.get("/api/settings/places-api-key")
async def places_api_key_status_api(session: Annotated[Session, Depends(get_session)]):
    from app.services.google_places_source import get_places_api_key

    key = get_places_api_key(session)
    return JSONResponse({"configured": bool(key), "masked": _mask_github_token(key)})


@router.post("/api/settings/places-api-key")
async def places_api_key_save_api(
    session: Annotated[Session, Depends(get_session)],
    key: str = Form(default=""),
):
    from app.services.google_places_source import save_places_api_key, get_places_api_key

    save_places_api_key(session, key)
    saved = get_places_api_key(session)
    return JSONResponse({"ok": True, "configured": bool(saved), "masked": _mask_github_token(saved)})


@router.get("/api/settings/places-usage")
async def places_usage_api(session: Annotated[Session, Depends(get_session)]):
    from app.services.google_places_source import get_usage

    return JSONResponse(get_usage(session))


async def _resolve_latest_release() -> dict:
    """Check GitHub releases for a newer version than the running app.
    Shared by the update-check poll and the start-update launcher below."""
    import requests as _requests

    current = settings.app_version
    headers = {"Accept": "application/vnd.github+json"}
    if settings.github_token:
        headers["Authorization"] = f"Bearer {settings.github_token}"

    try:
        resp = _requests.get(
            f"https://api.github.com/repos/{settings.update_repo}/releases/latest",
            headers=headers,
            timeout=5,
        )
        resp.raise_for_status()
        payload = resp.json()
        latest = str(payload.get("tag_name", "")).lstrip("v") or current
        release_url = payload.get(
            "html_url", f"https://github.com/{settings.update_repo}/releases/latest"
        )
        # Prefer a direct asset link over the release page so the update
        # button can start the download itself instead of sending the user
        # to GitHub to find and click the right file by hand.
        download_url = ""
        for asset in payload.get("assets", []) or []:
            name = str(asset.get("name", "")).lower()
            if name.endswith(".exe") and "setup" in name:
                download_url = asset.get("browser_download_url", "")
                break
        if not download_url:
            for asset in payload.get("assets", []) or []:
                if str(asset.get("name", "")).lower().endswith(".exe"):
                    download_url = asset.get("browser_download_url", "")
                    break
    except Exception:
        return {
            "current": current,
            "latest": current,
            "update_available": False,
            "release_url": "",
            "download_url": "",
        }

    def _version_tuple(value: str):
        parts = []
        for chunk in value.split("."):
            digits = "".join(ch for ch in chunk if ch.isdigit())
            parts.append(int(digits) if digits else 0)
        return tuple(parts)

    update_available = _version_tuple(latest) > _version_tuple(current)

    return {
        "current": current,
        "latest": latest,
        "update_available": update_available,
        "release_url": release_url,
        "download_url": download_url,
    }


@router.get("/api/app/update-check")
async def update_check_api():
    return JSONResponse(await _resolve_latest_release())


# In-process state for the installer download, polled by the settings panel
# so it can render a real progress bar instead of a single opaque "please
# wait". Single-slot (not keyed by session) — this app only ever runs one
# local instance per machine, so one in-flight download at a time is fine.
_update_download_state: dict = {
    "status": "idle",  # idle | downloading | done | error
    "percent": 0,
    "downloaded": 0,
    "total": 0,
    "error": "",
    "installer_path": "",
    "latest": "",
}
_update_download_lock = threading.Lock()


def _download_installer_worker(download_url: str, latest: str) -> None:
    import requests as _requests

    with _update_download_lock:
        _update_download_state.update(
            status="downloading", percent=0, downloaded=0, total=0, error="", installer_path="", latest=latest
        )
    try:
        resp = _requests.get(download_url, timeout=120, stream=True)
        resp.raise_for_status()
        total = int(resp.headers.get("content-length") or 0)
        installer_path = Path(tempfile.gettempdir()) / "JobMind-Match-Setup.exe"
        downloaded = 0
        with open(installer_path, "wb") as fh:
            for chunk in resp.iter_content(chunk_size=1 << 16):
                if not chunk:
                    continue
                fh.write(chunk)
                downloaded += len(chunk)
                percent = int(downloaded * 100 / total) if total else 0
                with _update_download_lock:
                    _update_download_state.update(status="downloading", percent=percent, downloaded=downloaded, total=total)
        with _update_download_lock:
            _update_download_state.update(status="done", percent=100, installer_path=str(installer_path))
    except Exception as exc:
        with _update_download_lock:
            _update_download_state.update(status="error", error=str(exc))


@router.post("/api/app/download-update")
async def download_update_api():
    """Kick off the installer download in a background thread. Windows-only:
    the release assets this resolves are .exe installers, and the install
    step (os.startfile / subprocess launch) only applies to Windows."""
    if platform.system() != "Windows":
        return JSONResponse(
            {"ok": False, "error": "Direct install is only available on Windows builds right now."},
            status_code=400,
        )

    with _update_download_lock:
        already_running = _update_download_state["status"] == "downloading"
    if already_running:
        return JSONResponse({"ok": True})

    data = await _resolve_latest_release()
    download_url = data.get("download_url", "")
    if not download_url:
        return JSONResponse(
            {"ok": False, "error": "Could not find an installer file for the latest release."},
            status_code=400,
        )

    threading.Thread(
        target=_download_installer_worker, args=(download_url, data.get("latest", "")), daemon=True
    ).start()
    return JSONResponse({"ok": True})


@router.get("/api/app/download-progress")
async def download_progress_api():
    with _update_download_lock:
        return JSONResponse(dict(_update_download_state))


def _resolve_close_target() -> tuple[int, str]:
    """Which process/.exe the update flow must confirm closed before ever
    touching the installer. Real packaged-install testing found this route
    actually executes inside a *uvicorn child process*, spawned by the real
    Windows launcher (scripts/desktop_launcher.py, a pywebview native-window
    app) via `subprocess.Popen([bundled_python, "-m", "uvicorn", ...])` —
    os.getpid()/sys.executable here name that child, never JobMindMatch.exe
    itself, which stays running (blocked in webview.start()) holding the
    real lock on its own .exe the whole time. That launcher now writes its
    own PID to launcher.pid at startup for exactly this reason (see
    write_launcher_pid/close_running_instance there) — prefer that target
    when present. Falls back to this process's own PID/exe otherwise, which
    is correct for the macOS/Linux onedir build (jobmind.spec + the root
    desktop_launcher.py), where the launcher and the server are the same
    single process, and is a safe default in any other environment lacking
    the marker (e.g. running the dev server directly)."""
    try:
        launcher_pid = int((app_root() / "launcher.pid").read_text(encoding="utf-8").strip())
        return launcher_pid, str(app_root() / "JobMindMatch.exe")
    except (OSError, ValueError):
        return os.getpid(), sys.executable


def _close_window_owned_by(pid: int) -> None:
    """PostMessage WM_CLOSE to the launcher's own "JobMind Match" pywebview
    window (ported from close_window_owned_by in scripts/desktop_launcher.py
    — same technique, needed here too because the wait-loop below only
    waits for that process to exit, it never asks it to. Without this, an
    update triggered from the dashboard would wait on the launcher's PID
    forever: nothing else in this flow tells that separate process to
    close its window, since (unlike the old single-process assumption)
    os._exit(0) a few lines down only ends the uvicorn child this route
    runs in, not the launcher."""
    if sys.platform != "win32":
        return
    try:
        WM_CLOSE = 0x0010
        user32 = ctypes.windll.user32

        @ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.c_void_p, ctypes.c_void_p)
        def _enum_cb(hwnd, _lparam):
            length = user32.GetWindowTextLengthW(hwnd)
            if length:
                buf = ctypes.create_unicode_buffer(length + 1)
                user32.GetWindowTextW(hwnd, buf, length + 1)
                if buf.value == "JobMind Match":
                    owner_pid = ctypes.c_ulong()
                    user32.GetWindowThreadProcessId(hwnd, ctypes.byref(owner_pid))
                    if owner_pid.value == pid:
                        user32.PostMessageW(hwnd, WM_CLOSE, 0, 0)
            return True

        user32.EnumWindows(_enum_cb, 0)
    except Exception:
        pass


def _spawn_update_after_current_process_exits(installer_path: str) -> None:
    """Ported from MessageCannon Pro's update_checker.py, where launching the
    installer and closing the app at essentially the same moment (Popen +
    a short sleep + os._exit) was verified via a real controlled
    reproduction to fail outright with Inno Setup exit code 5 — the
    installer can't overwrite this app's own locked, in-use .exe. The
    previous fire-and-forget Popen here had that same shape and never
    checked for that failure, so it would have looked identical to a
    successful update while silently leaving the user on the old version.

    A first hardening pass here (poll for the PID to disappear instead of a
    single fallible `Wait-Process` call, then separately poll for the .exe
    to become exclusively-openable before ever touching the installer —
    "confirmed closed", not "close requested") still weren't enough: see
    _resolve_close_target — this route runs inside a different process than
    the one that actually needs to be closed on the real Windows build.
    installer/jobmind-match.iss uses PrivilegesRequired=lowest like
    MessageCannon's, so the same silent, unelevated re-run is safe here too.
    `CREATE_NEW_PROCESS_GROUP` alone (no `DETACHED_PROCESS`) is required —
    MessageCannon found DETACHED_PROCESS reliably prevented the helper from
    completing its job."""
    pid, exe_path = _resolve_close_target()
    if pid != os.getpid():
        # Different process (the real launcher) -- ask it to close. When
        # pid == os.getpid() (no launcher.pid marker found), this route's
        # own os._exit(0) a moment later is what closes it instead; no
        # window to close for that case.
        _close_window_owned_by(pid)
    install_cmd = [installer_path, "/VERYSILENT", "/SUPPRESSMSGBOXES", "/NORESTART"]
    escaped_installer_path = install_cmd[0].replace("'", "''")
    escaped_exe_path = exe_path.replace("'", "''")
    args_literal = ",".join(f"'{a}'" for a in install_cmd[1:])
    ps_script = (
        f"$targetPid = {pid}\n"
        f"$exePath = '{escaped_exe_path}'\n"
        "$processDeadline = (Get-Date).AddSeconds(30)\n"
        "while ((Get-Date) -lt $processDeadline) {\n"
        "    $proc = Get-Process -Id $targetPid -ErrorAction SilentlyContinue\n"
        "    if (-not $proc) { break }\n"
        "    Start-Sleep -Milliseconds 150\n"
        "}\n"
        "$fileDeadline = (Get-Date).AddSeconds(10)\n"
        "while ((Get-Date) -lt $fileDeadline) {\n"
        "    try {\n"
        "        $stream = [System.IO.File]::Open($exePath, 'Open', 'ReadWrite', 'None')\n"
        "        $stream.Close()\n"
        "        break\n"
        "    } catch {\n"
        "        Start-Sleep -Milliseconds 200\n"
        "    }\n"
        "}\n"
        f"Start-Process -FilePath '{escaped_installer_path}' -ArgumentList {args_literal}\n"
    )
    command = ["powershell.exe", "-NoProfile", "-WindowStyle", "Hidden", "-Command", ps_script]
    kwargs: dict = {}
    if sys.platform == "win32":
        kwargs["creationflags"] = subprocess.CREATE_NEW_PROCESS_GROUP
    subprocess.Popen(
        command, close_fds=True,
        stdin=subprocess.DEVNULL, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
        **kwargs,
    )


@router.post("/api/app/install-update")
async def install_update_api():
    """Launch the already-downloaded installer once this app has fully
    closed, then close this app — see _spawn_update_after_current_process_exits."""
    with _update_download_lock:
        state = dict(_update_download_state)

    installer_path = state.get("installer_path", "")
    if state.get("status") != "done" or not installer_path or not Path(installer_path).exists():
        return JSONResponse({"ok": False, "error": "Installer not downloaded yet."}, status_code=400)

    try:
        _spawn_update_after_current_process_exits(installer_path)
    except Exception as exc:
        return JSONResponse({"ok": False, "error": f"Could not launch installer: {exc}"}, status_code=500)

    async def _shutdown_soon() -> None:
        # Give the JSON response time to actually reach the browser before
        # the process exits — the PowerShell helper above is what actually
        # waits for this exit to complete before touching the installer, so
        # this delay is purely for the HTTP response, not installer safety.
        await asyncio.sleep(0.8)
        os._exit(0)

    asyncio.create_task(_shutdown_soon())
    return JSONResponse({"ok": True})


@router.post("/api/app/quit")
async def quit_app_api():
    """Graceful-close target for desktop_launcher.py's `--quit` (the Start
    Menu "Quit" shortcut and the installer's [UninstallRun] step both invoke
    it). No installer to spawn here, so the response-flush delay can be
    short — the caller is the one polling for this process's PID to
    actually disappear, same "confirm closed" pattern as the update flow."""
    async def _shutdown_soon() -> None:
        await asyncio.sleep(0.3)
        os._exit(0)

    asyncio.create_task(_shutdown_soon())
    return JSONResponse({"ok": True})


@router.post("/api/scrape/export")
async def export_scraped_leads(
    request: Request,
    session: Annotated[Session, Depends(get_session)],
    fmt: str = Query(default="csv"),
    filename: str = Query(default="leads"),
    session_id: Optional[int] = Query(default=None),
):
    """Export scraped leads in multiple formats. Leads are normally posted as a
    JSON body — a hunt session can hold thousands of rows, well past what a
    URL can carry. If `session_id` is given, leads are instead read straight
    from disk (the HuntedContact table), so a recovered/past hunt can still be
    exported even if the page never held the array in memory (reload, crash,
    restart)."""
    try:
        payload = await request.json()
        leads = payload if isinstance(payload, list) else payload.get("leads", [])
    except Exception:
        leads = []

    if not leads and session_id is not None:
        from app.services.lead_hunter_registry import get_session_leads

        leads = get_session_leads(session, session_id)

    if not leads:
        raise HTTPException(status_code=400, detail="No leads to export")

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = re.sub(r'[^\w\-]', '_', filename) or 'leads'
    fname = f"{safe_name}_{ts}"

    if fmt == "csv":
        out = io.StringIO()
        w = csv.writer(out)
        w.writerow(['#','Name','Designation','Email','WhatsApp','Source','Profile URL','Notes'])
        for i, l in enumerate(leads, 1):
            w.writerow([i, l.get('name',''), l.get('designation',''), l.get('email',''), l.get('whatsapp',''), l.get('source',''), l.get('url',''), l.get('notes','')])
        out.seek(0)
        return StreamingResponse(
            iter([out.getvalue()]),
            media_type='text/csv',
            headers={'Content-Disposition': f'attachment; filename="{fname}.csv"'}
        )

    elif fmt == "json":
        return StreamingResponse(
            iter([json.dumps(leads, indent=2, ensure_ascii=False)]),
            media_type='application/json',
            headers={'Content-Disposition': f'attachment; filename="{fname}.json"'}
        )

    elif fmt == "txt_email":
        seen: set[str] = set()
        emails: list[str] = []
        for l in leads:
            email = (l.get('email') or '').strip()
            # "low" = matches an obvious placeholder/example pattern
            # (is_placeholder_email) -- excluded so this list stays safe to
            # hand off for real outreach.
            if not email or l.get('confidence') == 'low':
                continue
            key = email.lower()
            if key in seen:
                continue
            seen.add(key)
            emails.append(email)
        return StreamingResponse(
            iter(['\n'.join(emails)]),
            media_type='text/plain',
            headers={'Content-Disposition': f'attachment; filename="{fname}_emails.txt"'}
        )

    elif fmt == "txt_wa":
        seen_wa: set[str] = set()
        wa_list: list[str] = []
        for l in leads:
            wa = (l.get('whatsapp') or '').strip()
            if not wa or l.get('confidence') == 'low' or not _E164_RE.match(wa):
                continue
            if wa in seen_wa:
                continue
            seen_wa.add(wa)
            wa_list.append(wa)
        return StreamingResponse(
            iter(['\n'.join(wa_list)]),
            media_type='text/plain',
            headers={'Content-Disposition': f'attachment; filename="{fname}_whatsapp.txt"'}
        )

    elif fmt in ("xls", "xlsx"):
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        ws = wb.active
        ws.title = "Leads"

        headers = ['#', 'Name', 'Designation', 'Email', 'WhatsApp', 'Source', 'Profile URL', 'Notes']
        ws.append(headers)
        header_fill = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")

        for i, l in enumerate(leads, 1):
            ws.append([
                i,
                l.get('name', ''),
                l.get('designation', ''),
                l.get('email', ''),
                l.get('whatsapp', ''),
                l.get('source', ''),
                l.get('url', ''),
                l.get('notes', ''),
            ])

        widths = [4, 22, 26, 28, 18, 16, 40, 40]
        for idx, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = width
        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{fname}.xlsx"'}
        )

    elif fmt == "html":
        cards_parts = []
        for i, l in enumerate(leads, 1):
            email_html = ""
            if l.get("email"):
                email = l.get("email", "")
                email_html = (
                    f"<a href='mailto:{email}' style='color:#60a5fa;font-size:12px'>"
                    f"✉ {email}</a><br>"
                )
            wa_html = ""
            if l.get("whatsapp"):
                wa = l.get("whatsapp", "")
                wa_digits = wa.replace("+", "")
                wa_html = (
                    f"<a href='https://wa.me/{wa_digits}' style='color:#25d366;font-size:12px'>"
                    f"💬 {wa}</a><br>"
                )
            cards_parts.append(
                f'<div style="border:1px solid #7c3aed;border-radius:10px;padding:16px;margin:8px;'
                f'display:inline-block;min-width:240px;max-width:300px;background:#0e0e1c;color:#eeeef8;'
                f'vertical-align:top;font-family:sans-serif">'
                f'<div style="font-weight:700;font-size:14px;margin-bottom:4px">#{i} {l.get("name", "Unknown")}</div>'
                f'<div style="font-size:11px;color:#a07cfc;margin-bottom:8px">{l.get("designation", "")}</div>'
                f'{email_html}{wa_html}'
                f'<div style="font-size:10px;color:#6666a0;margin-top:6px">{l.get("source", "")}</div>'
                f'</div>'
            )
        cards = "".join(cards_parts)
        html_out = (
            f'<!DOCTYPE html><html><head><meta charset="UTF-8">'
            f'<title>Lead Hunter — {len(leads)} Contacts</title>'
            f'<style>body{{background:#080812;padding:24px;font-family:sans-serif}}'
            f'h2{{color:#c084fc;margin-bottom:20px}}'
            f'.stats{{color:#6666a0;font-size:13px;margin-bottom:20px}}</style></head>'
            f'<body><h2>🎯 Lead Hunter Pro — {len(leads)} Contacts</h2>'
            f'<div class="stats">Generated: {datetime.now().strftime("%d %b %Y %H:%M")} | '
            f'Emails: {sum(1 for l in leads if l.get("email"))} | '
            f'WhatsApp: {sum(1 for l in leads if l.get("whatsapp"))}</div>'
            f'{cards}</body></html>'
        )
        return StreamingResponse(
            iter([html_out]),
            media_type='text/html',
            headers={'Content-Disposition': f'attachment; filename="{fname}.html"'}
        )

    raise HTTPException(status_code=400, detail=f"Invalid format: {fmt}. Use: csv, json, xls, xlsx, txt_email, txt_wa, html")